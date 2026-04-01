"""TB_Wave - 수동 밸런싱, 가변 웨이브 간격"""
import openpyxl

wb = openpyxl.load_workbook('car_survivor_tables.xlsx')
ws = wb['TB_Wave']

headers = ['wave_group_id', 'wave_no', 'mon_id', 'spawn_count', 'spawn_interval', 'max_enemies', 'difficulty_scale', 'note']
for i, h in enumerate(headers):
    ws.cell(row=1, column=i + 1, value=h)
for col in range(len(headers) + 1, 15):
    ws.cell(row=1, column=col, value=None)

for row in range(2, ws.max_row + 1):
    for col in range(1, 15):
        ws.cell(row=row, column=col, value=None)

# (wave_group_id, wave_no, mon_id, spawn_count, spawn_interval, max_enemies, difficulty_scale, note_base)
# spawn_count = 6초당 몇 마리, spawn_interval = 6 고정
waves = [
    # === 0~2분: 20초 간격 (Wave 1~6) ===
    # 0~20초
    ('WG_CH1', 1, 'MON_001', 2, 6, 15, 1.0, ''),
    # 20~40초
    ('WG_CH1', 2, 'MON_001', 3, 6, 18, 1.05, ''),
    # 40~60초
    ('WG_CH1', 3, 'MON_001', 2, 6, 20, 1.1, 'MON_002 첫 등장'),
    ('WG_CH1', 3, 'MON_002', 1, 6, 20, 1.1, ''),
    # 60~80초
    ('WG_CH1', 4, 'MON_001', 1, 6, 22, 1.15, ''),
    ('WG_CH1', 4, 'MON_002', 2, 6, 22, 1.15, ''),
    # 80~100초
    ('WG_CH1', 5, 'MON_001', 2, 6, 25, 1.2, 'MON_003 첫 등장'),
    ('WG_CH1', 5, 'MON_003', 2, 6, 25, 1.2, ''),
    # 100~120초
    ('WG_CH1', 6, 'MON_002', 4, 6, 28, 1.25, 'MON_002 러시'),

    # === 2~4분: 30초 간격 (Wave 7~10) ===
    # 120~150초
    ('WG_CH1', 7, 'MON_002', 3, 6, 30, 1.3, ''),
    ('WG_CH1', 7, 'MON_003', 1, 6, 30, 1.3, ''),
    # 150~180초
    ('WG_CH1', 8, 'MON_002', 2, 6, 32, 1.4, ''),
    ('WG_CH1', 8, 'MON_002', 2, 6, 32, 1.4, ''),
    # 180~210초
    ('WG_CH1', 9, 'MON_002', 2, 6, 35, 1.5, ''),
    ('WG_CH1', 9, 'MON_002', 2, 6, 35, 1.5, ''),
    # 210~240초
    ('WG_CH1', 10, 'MON_002', 3, 6, 38, 1.6, ''),
    ('WG_CH1', 10, 'MON_002', 2, 6, 38, 1.6, ''),

    # === 4~10분: 40초 간격 (Wave 11~19) ===
    # 240~280초
    ('WG_CH1', 11, 'MON_002', 3, 6, 40, 1.7, ''),
    ('WG_CH1', 11, 'MON_002', 3, 6, 40, 1.7, ''),
    # 280~320초
    ('WG_CH1', 12, 'MON_002', 4, 6, 42, 1.8, ''),
    ('WG_CH1', 12, 'MON_002', 3, 6, 42, 1.8, ''),
    # 320~360초 (빠진 구간 추가)
    ('WG_CH1', 13, 'MON_002', 4, 6, 45, 1.9, ''),
    ('WG_CH1', 13, 'MON_002', 3, 6, 45, 1.9, ''),
    # 360~400초
    ('WG_CH1', 14, 'MON_002', 4, 6, 48, 2.0, ''),
    ('WG_CH1', 14, 'MON_002', 4, 6, 48, 2.0, ''),
    # 400~440초
    ('WG_CH1', 15, 'MON_002', 3, 6, 50, 2.1, 'MON_003 재등장'),
    ('WG_CH1', 15, 'MON_002', 4, 6, 50, 2.1, ''),
    ('WG_CH1', 15, 'MON_003', 1, 6, 50, 2.1, ''),
    # 440~480초
    ('WG_CH1', 16, 'MON_002', 2, 6, 52, 2.2, ''),
    ('WG_CH1', 16, 'MON_002', 4, 6, 52, 2.2, ''),
    ('WG_CH1', 16, 'MON_003', 2, 6, 52, 2.2, ''),
    # 480~520초
    ('WG_CH1', 17, 'MON_002', 4, 6, 55, 2.4, ''),
    ('WG_CH1', 17, 'MON_003', 3, 6, 55, 2.4, ''),
    # 520~560초
    ('WG_CH1', 18, 'MON_002', 4, 6, 58, 2.6, ''),
    ('WG_CH1', 18, 'MON_003', 4, 6, 58, 2.6, ''),
    # 560~600초
    ('WG_CH1', 19, 'MON_002', 2, 6, 60, 3.0, '최종'),
    ('WG_CH1', 19, 'MON_003', 6, 6, 60, 3.0, ''),
]

# 웨이브 시작 시간 계산 (가변 간격)
wave_start_times = {}
# Wave 1~6: 20초 간격, 0초 시작
for i in range(1, 7):
    wave_start_times[i] = (i - 1) * 20
# Wave 7~10: 30초 간격, 120초 시작
for i in range(7, 11):
    wave_start_times[i] = 120 + (i - 7) * 30
# Wave 11~19: 40초 간격, 240초 시작
for i in range(11, 20):
    wave_start_times[i] = 240 + (i - 11) * 40

# note에 시간 주석 추가
seen = set()
for i, w in enumerate(waves):
    no = w[1]
    start = wave_start_times.get(no, 0)
    # 다음 웨이브 시작 = 현재 구간 끝
    next_nos = [n for n in wave_start_times if n > no]
    end = wave_start_times[min(next_nos)] if next_nos else 600
    time_str = f'{start // 60:02d}:{start % 60:02d}~{end // 60:02d}:{end % 60:02d}'
    note = w[7]
    if no not in seen:
        seen.add(no)
        if note:
            note = f'[{time_str}] {note}'
        else:
            note = f'[{time_str}]'
    waves[i] = w[:7] + (note,)

for i, w in enumerate(waves):
    row = i + 2
    for col, val in enumerate(w):
        ws.cell(row=row, column=col + 1, value=val)

if ws.max_row > len(waves) + 1:
    ws.delete_rows(len(waves) + 2, ws.max_row - len(waves) - 1)

wb.save('car_survivor_tables.xlsx')

unique_nos = sorted(set(w[1] for w in waves))
print(f'Done! {len(waves)} rows, {len(unique_nos)} waves')
for no in unique_nos:
    entries = [w for w in waves if w[1] == no]
    start = wave_start_times.get(no, 0)
    next_nos = [n for n in wave_start_times if n > no]
    end = wave_start_times[min(next_nos)] if next_nos else 600
    desc = ' + '.join(f'{w[2]}(x{w[3]}/{w[4]}s)' for w in entries)
    print(f'  Wave {no:2d} | {start//60:02d}:{start%60:02d}~{end//60:02d}:{end%60:02d} | {desc}')
