"""TB_Lang_LevelUpSelect 언어 테이블 생성 + Weapon/SpellBook에 lang_name, lang_desc 추가"""
import openpyxl

wb = openpyxl.load_workbook('car_survivor_tables.xlsx')

# === 1. TB_Weapon에 lang_name, lang_desc 추가 ===
ws_w = wb['TB_Weapon']
# 현재 데이터 읽기
weapon_data = []
for i, row in enumerate(ws_w.iter_rows(values_only=True)):
    if i == 0: continue
    if row[0] is None or str(row[0]).startswith('['): break
    weapon_data.append(list(row))

# 기존 weapon_name, effect_desc → lang_name, lang_desc 키로 변환
weapon_lang = {}
for w in weapon_data:
    wid = w[0]
    name = w[1] if w[1] else ''
    desc = w[4] if w[4] else ''
    lang_name_key = f'LANG_{wid}_NAME'
    lang_desc_key = f'LANG_{wid}_DESC'
    weapon_lang[lang_name_key] = name
    weapon_lang[lang_desc_key] = desc
    w[1] = lang_name_key   # weapon_name → lang key
    w[4] = lang_desc_key   # effect_desc → lang key

# 덮어쓰기
for i, w in enumerate(weapon_data):
    row = i + 2
    for col, val in enumerate(w):
        if col < ws_w.max_column:
            ws_w.cell(row=row, column=col + 1, value=val)

# === 2. TB_SpellBook에 lang_name, lang_desc 추가 ===
ws_s = wb['TB_SpellBook']
spell_data = []
for i, row in enumerate(ws_s.iter_rows(values_only=True)):
    if i == 0: continue
    if row[0] is None or str(row[0]).startswith('['): break
    spell_data.append(list(row))

spell_lang = {}
for s in spell_data:
    sid = s[0]
    name = s[1] if s[1] else ''
    desc = s[4] if s[4] else ''
    lang_name_key = f'LANG_{sid}_NAME'
    lang_desc_key = f'LANG_{sid}_DESC'
    spell_lang[lang_name_key] = name
    spell_lang[lang_desc_key] = desc
    s[1] = lang_name_key
    s[4] = lang_desc_key

for i, s in enumerate(spell_data):
    row = i + 2
    for col, val in enumerate(s):
        if col < ws_s.max_column:
            ws_s.cell(row=row, column=col + 1, value=val)

# === 3. TB_Lang_LevelUpSelect 시트 생성 ===
if 'TB_Lang_LevelUpSelect' in wb.sheetnames:
    del wb['TB_Lang_LevelUpSelect']

ws_lang = wb.create_sheet('TB_Lang_LevelUpSelect')
ws_lang.cell(row=1, column=1, value='lang_key')
ws_lang.cell(row=1, column=2, value='ko')
ws_lang.cell(row=1, column=3, value='en')

row = 2
# 무기
for key, ko_text in weapon_lang.items():
    ws_lang.cell(row=row, column=1, value=key)
    ws_lang.cell(row=row, column=2, value=ko_text)
    ws_lang.cell(row=row, column=3, value='')  # 영어는 나중에
    row += 1

# 스펠북
for key, ko_text in spell_lang.items():
    ws_lang.cell(row=row, column=1, value=key)
    ws_lang.cell(row=row, column=2, value=ko_text)
    ws_lang.cell(row=row, column=3, value='')
    row += 1

wb.save('car_survivor_tables.xlsx')

print(f'Done!')
print(f'  TB_Weapon: {len(weapon_data)} rows updated')
print(f'  TB_SpellBook: {len(spell_data)} rows updated')
print(f'  TB_Lang_LevelUpSelect: {row - 2} lang entries')
print()
print('=== 언어 테이블 내용 ===')
for key, ko in {**weapon_lang, **spell_lang}.items():
    print(f'  {key} → {ko}')
