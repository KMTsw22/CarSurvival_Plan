"""
엑셀 테이블 → MessagePack .bytes 변환 스크립트

사용법:
    py export_msgpack.py

입력: car_survivor_tables.xlsx (또는 _v2)
출력: ../../CarSurvior/Assets/Resources/Tables/*.bytes
"""

import sys
import io
import os
import msgpack
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ─── 설정 ───
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_CANDIDATES = [
    os.path.join(SCRIPT_DIR, "car_survivor_tables_v2.xlsx"),
    os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx"),
]
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

# 각 시트에 대한 설정: (시트명, 출력파일명, 컬럼 타입 리스트)
# 타입: str, int, float, bool
SHEET_CONFIG = {
    "TB_Currency": {
        "output": "TB_Currency",
        "types": [str, str, str, str],
    },
    "TB_Car": {
        "output": "TB_Car",
        "types": [str, str, str, float, float, float, float, float, str, float, str, int, str, bool, str, float, float],
    },
    "TB_Weapon": {
        "output": "TB_Weapon",
        "types": [str, str, str, float, str, str, str, float, float, int, int, str, float, float, float, float, float, float],
    },
    "TB_SpellBook": {
        "output": "TB_SpellBook",
        "types": [str, str, str, float, str, int, bool, int, str],
    },
    "TB_Part": {
        "output": "TB_Part",
        "types": [str, str, str, str, float, int, str],
    },
    "TB_Monster": {
        "output": "TB_Monster",
        "types": [str, str, bool, float, float, float, float, int, int, str, str, float, float, float],
    },
    "TB_MonsterDrop": {
        "output": "TB_MonsterDrop",
        "types": [str, str, int, int, int],
    },
    "TB_Wave": {
        "output": "TB_Wave",
        "types": [str, int, str, int, float, int, float, str, float, float, float, int, float, float, float],
    },
    "TB_Level": {
        "output": "TB_Level",
        "types": [int, int, int, float, str],
    },
    "TB_LangLevelUpSelect_name": {
        "output": "TB_LangLevelUpSelect_name",
        "types": [str, str, str],
    },
    "TB_LangLevelUpSelect_des": {
        "output": "TB_LangLevelUpSelect_des",
        "types": [str, str, str],
    },
    "TB_Map": {
        "output": "TB_Map",
        "types": [str, str, str, int, str, str, int, str, bool],
    },
    "TB_Stage": {
        "output": "TB_Stage",
        "types": [str, str, int, str, str, int, str, str, float, str],
    },
    "TB_WarningWave": {
        "output": "TB_WarningWave",
        "types": [str, int, str, int, float, int, float, float, str],
    },
}


def cast_value(value, target_type):
    """셀 값을 지정된 타입으로 변환"""
    if value is None:
        if target_type == str:
            return ""
        elif target_type == int:
            return 0
        elif target_type == float:
            return 0.0
        elif target_type == bool:
            return False
        return value

    if target_type == bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ("true", "1", "yes")
        return bool(value)
    elif target_type == int:
        return int(float(value)) if value else 0
    elif target_type == float:
        return float(value) if value else 0.0
    elif target_type == str:
        return str(value) if value else ""
    return value


def process_sheet(ws, config):
    """시트 하나를 읽어서 MessagePack 배열로 변환"""
    types = config["types"]
    rows = []
    header_count = len(types)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # 데이터 영역만 읽기 (빈 행이나 설명 섹션 건너뛰기)
        first_cell = row[0] if row else None
        if first_cell is None or str(first_cell).startswith("["):
            # 빈 행이거나 설명 섹션([컬럼명]) 시작 → 중단
            break

        row_data = []
        for col_idx in range(header_count):
            raw = row[col_idx] if col_idx < len(row) else None
            row_data.append(cast_value(raw, types[col_idx]))

        rows.append(row_data)

    return rows


def main():
    # 엑셀 파일 찾기
    xlsx_path = None
    for candidate in XLSX_CANDIDATES:
        if os.path.exists(candidate):
            xlsx_path = candidate
            break

    if xlsx_path is None:
        print("ERROR: car_survivor_tables.xlsx 파일을 찾을 수 없습니다.")
        sys.exit(1)

    print("Input: {}".format(xlsx_path))

    # 출력 폴더 생성
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Output: {}".format(os.path.abspath(OUTPUT_DIR)))

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    total = 0
    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print("  SKIP: {} (시트 없음)".format(sheet_name))
            continue

        ws = wb[sheet_name]
        rows = process_sheet(ws, config)

        # MessagePack 직렬화
        packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)

        # .bytes 파일 저장
        out_path = os.path.join(OUTPUT_DIR, config["output"] + ".bytes")
        with open(out_path, "wb") as f:
            f.write(packed)

        print("  {} -> {} ({} rows, {} bytes)".format(
            sheet_name, os.path.basename(out_path), len(rows), len(packed)))
        total += len(rows)

    wb.close()
    print("\nDone! Total {} rows exported to {} tables.".format(total, len(SHEET_CONFIG)))


if __name__ == "__main__":
    main()
