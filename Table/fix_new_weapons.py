"""
TB_Weapon: 새 무기 이름/설명 영어로 변경 + etc_value 설명 섹션 갱신 + .bytes 재익스포트

사용법: py fix_new_weapons.py
주의: 엑셀 파일을 먼저 닫고 실행하세요!
"""
import os, sys, io, msgpack, openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

TYPES = [str, str, str, float, str, str, str, float, float, int, int, str,
         float, float, float, float, float, float]

# weapon_id → 영어로 변경할 데이터 (weapon_name, effect_desc)
EN_DATA = {
    "WPN_004": ("Chain Lightning", "Chain lightning attack"),
    "WPN_005": ("EMP Pulse",       "Area electromagnetic shock"),
    "WPN_006": ("Flamethrower",    "Frontal flame spray"),
    "WPN_007": ("Laser Cannon",    "Piercing laser beam"),
    "WPN_008": ("Missile Pod",     "Homing missile salvo"),
}

# 컬럼 설명 섹션 (기존 + 신규 포함, 18컬럼)
DESCRIPTIONS = [
    ("weapon_id",        "string (PK)", "Unique weapon ID. Format: WPN_###"),
    ("weapon_name",      "string",      "Weapon display name"),
    ("weapon_category",  "enum",        "Main = primary / Sub = secondary"),
    ("damage",           "float",       "Base damage (Lv.1)"),
    ("effect_desc",      "string",      "Effect description"),
    ("aim_type",         "enum",        "Manual = mouse aim / Auto = auto trigger"),
    ("weapon_type",      "enum",        "MachineGun / OilSlick / SawBlade / ChainLightning / EMPPulse / Flamethrower / LaserCannon / MissilePod"),
    ("cooldown",         "float",       "Cooldown / fire interval (sec). 0 = none"),
    ("duration",         "float",       "Duration (sec). 0 = permanent"),
    ("max_level",        "int",         "Max upgrade level"),
    ("drop_weight",      "int",         "Level-up selection weight"),
    ("icon_key",         "string",      "Icon resource key"),
    ("etc_value1",       "float",       "MachineGun: (unused)\n"
                                        "OilSlick: slow % (e.g. 50 = 50%)\n"
                                        "SawBlade: rotation speed (deg/sec)\n"
                                        "ChainLightning: base chain target count\n"
                                        "EMPPulse: base stun duration (sec)\n"
                                        "Flamethrower: base flame duration (sec)\n"
                                        "LaserCannon: base laser count\n"
                                        "MissilePod: base missile count"),
    ("etc_value2",       "float",       "MachineGun: (unused)\n"
                                        "OilSlick: slow duration (sec)\n"
                                        "SawBlade: orbit radius\n"
                                        "ChainLightning: chain targets added per level\n"
                                        "EMPPulse: base radius\n"
                                        "Flamethrower: base radius\n"
                                        "LaserCannon: lasers added per level\n"
                                        "MissilePod: missiles added per level"),
    ("etc_value3",       "float",       "MachineGun: (unused)\n"
                                        "OilSlick: base puddle radius\n"
                                        "SawBlade: hit interval (sec)\n"
                                        "ChainLightning: (unused)\n"
                                        "EMPPulse: radius increase per level\n"
                                        "Flamethrower: radius increase per level\n"
                                        "LaserCannon: (unused)\n"
                                        "MissilePod: (unused)"),
    ("etc_value4",       "float",       "Reserved (per-weapon extra param)"),
    ("etc_value5",       "float",       "Reserved (per-weapon extra param)"),
    ("damage_per_level", "float",       "Damage increase per level"),
]

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
desc_header_font = Font(bold=True, color='FFFFFF', size=10)
desc_header_fill = PatternFill(start_color='607D8B', end_color='607D8B', fill_type='solid')
desc_font = Font(color='37474F', size=10)
desc_fill = PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid')


def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    # --- 1) 무기 이름/설명 영어로 변경 ---
    last_data_row = 1
    for row in ws.iter_rows(min_row=2):
        cell_id = row[0].value
        if cell_id is None or str(cell_id).startswith("["):
            break
        last_data_row = row[0].row

        if cell_id in EN_DATA:
            name, desc = EN_DATA[cell_id]
            ws.cell(row=row[0].row, column=2).value = name   # weapon_name
            ws.cell(row=row[0].row, column=5).value = desc   # effect_desc
            print(f"  [EN] {cell_id}: {name} / {desc}")

    # --- 2) 기존 설명 섹션 삭제 ---
    # 데이터 영역 이후의 모든 행 클리어
    for row_num in range(last_data_row + 1, ws.max_row + 1):
        for col_num in range(1, 20):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()

    # --- 3) 새 설명 섹션 작성 ---
    desc_start = last_data_row + 2  # 빈 행 1줄 띄우고

    # 섹션 헤더
    for ci, label in enumerate(["[Column]", "[Type]", "[Description]"]):
        cell = ws.cell(row=desc_start, column=ci + 1, value=label)
        cell.font = desc_header_font
        cell.fill = desc_header_fill
        cell.border = thin_border

    # 각 컬럼 설명
    for i, (col_name, col_type, col_desc) in enumerate(DESCRIPTIONS):
        r = desc_start + 1 + i
        c1 = ws.cell(row=r, column=1, value=col_name)
        c1.font = Font(bold=True, color='37474F', size=10)
        c1.fill = desc_fill
        c1.border = thin_border

        c2 = ws.cell(row=r, column=2, value=col_type)
        c2.font = desc_font
        c2.fill = desc_fill
        c2.border = thin_border

        c3 = ws.cell(row=r, column=3, value=col_desc)
        c3.font = desc_font
        c3.fill = desc_fill
        c3.border = thin_border
        c3.alignment = Alignment(wrap_text=True, vertical='top')

    print(f"  [+] 설명 섹션 갱신 완료 ({len(DESCRIPTIONS)}개 컬럼)")

    # 저장
    wb.save(XLSX_PATH)
    wb.close()
    print(f"엑셀 저장 완료: {XLSX_PATH}")

    # --- 4) .bytes 재익스포트 ---
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_Weapon"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["):
            break
        row_data = [cast(row[i] if i < len(row) else None, TYPES[i]) for i in range(len(TYPES))]
        rows.append(row_data)
    wb.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_Weapon.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)
    print(f"TB_Weapon.bytes: {len(rows)} rows, {len(packed)} bytes")


if __name__ == "__main__":
    main()
