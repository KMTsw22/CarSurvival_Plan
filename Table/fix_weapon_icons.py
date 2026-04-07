"""
TB_Weapon icon_key 수정 스크립트
- Weapons 폴더의 실제 파일명을 읽어서 weapon_type과 매칭
- Excel의 icon_key 컬럼을 업데이트
- TB_Weapon.bytes 재출력

사용법: py fix_weapon_icons.py
"""

import os
import sys
import io
import msgpack
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
ICONS_DIR = os.path.join(
    SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources",
    "Sprites", "Icons", "Weapons"
)
OUTPUT_DIRS = [
    os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables"),
    os.path.join(SCRIPT_DIR, "..", "..", "CarSurvival", "Assets", "Resources", "Tables"),
]

WEAPON_TYPE_COL = 7   # 1-based (column G)
ICON_KEY_COL = 12     # 1-based (column L)

# Export types for TB_Weapon (18 columns)
TYPES = [str, str, str, float, str, str, str, float, float, int, int, str,
         float, float, float, float, float, float]


def cast_value(value, target_type):
    if value is None:
        if target_type == str:
            return ""
        elif target_type == int:
            return 0
        elif target_type == float:
            return 0.0
        elif target_type == bool:
            return False
        return value
    if target_type == bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ("true", "1", "yes")
        return bool(value)
    elif target_type == int:
        return int(float(value)) if value else 0
    elif target_type == float:
        return float(value) if value else 0.0
    elif target_type == str:
        return str(value) if value else ""
    return value


def main():
    # 1) Read actual icon filenames (without extension, excluding .meta)
    icon_files = []
    for f in os.listdir(ICONS_DIR):
        if f.endswith(".meta"):
            continue
        name_no_ext = os.path.splitext(f)[0]
        icon_files.append(name_no_ext)

    print("=== Actual icon files ===")
    for f in sorted(icon_files):
        print(f"  {f}")

    # 2) Build mapping: suffix after "ico-" -> full icon name
    #    e.g. "MachineGun" -> "ico-MachineGun"
    icon_name_map = {}
    for icon in icon_files:
        if icon.startswith("ico-"):
            suffix = icon[4:]
            icon_name_map[suffix] = icon

    print(f"\n=== Icon name map (weapon_type suffix -> icon_key) ===")
    for k, v in sorted(icon_name_map.items()):
        print(f"  {k} -> {v}")

    # 3) Open Excel and update icon_key based on weapon_type
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["TB_Weapon"]

    print(f"\n=== Updating icon_key values ===")
    changes = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        weapon_id = row[0].value
        if weapon_id is None:
            break

        weapon_type = row[WEAPON_TYPE_COL - 1].value
        old_icon_key = row[ICON_KEY_COL - 1].value

        # Match weapon_type to icon file
        if weapon_type in icon_name_map:
            new_icon_key = icon_name_map[weapon_type]
        else:
            print(f"  WARNING: No icon file for weapon_type '{weapon_type}', keeping '{old_icon_key}'")
            continue

        if old_icon_key != new_icon_key:
            print(f"  {weapon_id} ({weapon_type}): '{old_icon_key}' -> '{new_icon_key}'")
            row[ICON_KEY_COL - 1].value = new_icon_key
            changes += 1
        else:
            print(f"  {weapon_id} ({weapon_type}): '{old_icon_key}' (unchanged)")

    if changes > 0:
        wb.save(EXCEL_PATH)
        print(f"\nExcel saved with {changes} change(s).")
    else:
        wb.save(EXCEL_PATH)
        print(f"\nNo changes needed in Excel (saved anyway for consistency).")
    wb.close()

    # 4) Re-export TB_Weapon.bytes
    print(f"\n=== Re-exporting TB_Weapon.bytes ===")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws_ro = wb["TB_Weapon"]

    rows = []
    header_count = len(TYPES)
    for row in ws_ro.iter_rows(min_row=2, values_only=True):
        first_cell = row[0] if row else None
        if first_cell is None or str(first_cell).startswith("["):
            break
        row_data = []
        for col_idx in range(header_count):
            raw = row[col_idx] if col_idx < len(row) else None
            row_data.append(cast_value(raw, TYPES[col_idx]))
        rows.append(row_data)
    wb.close()

    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)

    for out_dir in OUTPUT_DIRS:
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, "TB_Weapon.bytes")
        with open(out_path, "wb") as f:
            f.write(packed)
        print(f"  Written: {os.path.abspath(out_path)} ({len(rows)} rows, {len(packed)} bytes)")

    print("\nDone!")


if __name__ == "__main__":
    main()
