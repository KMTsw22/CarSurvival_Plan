"""
TB_SpellBook에 공격력 증가 스펠북 추가 후 재익스포트

사용법: py add_atkup_spellbook.py
"""
import os
import sys
import io
import msgpack
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

# book_id, book_name, effect_type, base_value, effect_desc, max_level, stackable, drop_weight, icon_key
NEW_ROW = ["BOOK_ATKUP", "공격력 증가", "DamageUp", 10.0, "공격력 10% 증가", 5, True, 10, "ico_atkup"]

TYPES = [str, str, str, float, str, int, bool, int, str]

def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == bool:
        if isinstance(val, bool): return val
        if isinstance(val, str): return val.lower() in ("true", "1", "yes")
        return bool(val)
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_SpellBook"]

    # 중복 체크
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == "BOOK_ATKUP":
            print("BOOK_ATKUP 이미 존재합니다. 스킵.")
            wb.close()
            # 그래도 bytes는 재생성
            break
    else:
        # 마지막 행에 추가
        next_row = ws.max_row + 1
        for col, val in enumerate(NEW_ROW, 1):
            ws.cell(row=next_row, column=col, value=val)
        wb.save(XLSX_PATH)
        print(f"BOOK_ATKUP 추가 완료 (행 {next_row})")

    wb.close()

    # 전체 TB_SpellBook 재익스포트
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_SpellBook"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["):
            break
        converted = [cast(row[i] if i < len(row) else None, TYPES[i]) for i in range(len(TYPES))]
        rows.append(converted)

    wb.close()

    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_SpellBook.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)

    print(f"TB_SpellBook.bytes 재생성 완료: {len(rows)} rows, {len(packed)} bytes")

    # 언어 테이블에도 추가
    wb = openpyxl.load_workbook(XLSX_PATH)

    # 이름
    ws_name = wb["TB_LangLevelUpSelect_name"]
    found_name = False
    for row in ws_name.iter_rows(min_row=2, values_only=True):
        if row and row[0] == "BOOK_ATKUP":
            found_name = True
            break
    if not found_name:
        nr = ws_name.max_row + 1
        ws_name.cell(row=nr, column=1, value="BOOK_ATKUP")
        ws_name.cell(row=nr, column=2, value="공격력 증가")
        ws_name.cell(row=nr, column=3, value="Attack Up")

    # 설명
    ws_des = wb["TB_LangLevelUpSelect_des"]
    found_des = False
    for row in ws_des.iter_rows(min_row=2, values_only=True):
        if row and row[0] == "BOOK_ATKUP":
            found_des = True
            break
    if not found_des:
        nr = ws_des.max_row + 1
        ws_des.cell(row=nr, column=1, value="BOOK_ATKUP")
        ws_des.cell(row=nr, column=2, value="공격력 10% 증가")
        ws_des.cell(row=nr, column=3, value="Increase attack by 10%")

    wb.save(XLSX_PATH)
    wb.close()

    # 언어 테이블 재익스포트
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    for sheet, out_name, types in [
        ("TB_LangLevelUpSelect_name", "TB_LangLevelUpSelect_name", [str, str, str]),
        ("TB_LangLevelUpSelect_des", "TB_LangLevelUpSelect_des", [str, str, str]),
    ]:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or str(row[0]).startswith("["):
                break
            converted = [cast(row[i] if i < len(row) else None, types[i]) for i in range(len(types))]
            rows.append(converted)

        packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
        out_path = os.path.join(OUTPUT_DIR, out_name + ".bytes")
        with open(out_path, "wb") as f:
            f.write(packed)
        print(f"{out_name}.bytes 재생성 완료: {len(rows)} rows")

    wb.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
