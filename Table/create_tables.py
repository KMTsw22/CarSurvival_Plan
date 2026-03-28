import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === Style ===
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
pk_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
fk_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
enum_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
desc_header_font = Font(bold=True, color='FFFFFF', size=10)
desc_header_fill = PatternFill(start_color='607D8B', end_color='607D8B', fill_type='solid')
desc_font = Font(color='37474F', size=10)
desc_fill = PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def add_data(ws, headers, data, col_types=None, descriptions=None):
    # Header row
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    # Data rows
    for r in data:
        ws.append(r)
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=1 + len(data), max_col=len(headers)):
        for col_idx, cell in enumerate(row, 1):
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_types and col_idx - 1 < len(col_types):
                t = col_types[col_idx - 1]
                if t == 'pk':
                    cell.fill = pk_fill
                elif t == 'fk':
                    cell.fill = fk_fill
                elif t == 'enum':
                    cell.fill = enum_fill
    # Column descriptions at the bottom
    if descriptions:
        desc_start = 1 + len(data) + 2  # 1 header + data + 1 blank row
        # Section header
        label_cell = ws.cell(row=desc_start, column=1, value='[컬럼명]')
        label_cell.font = desc_header_font
        label_cell.fill = desc_header_fill
        label_cell.border = thin_border
        type_cell = ws.cell(row=desc_start, column=2, value='[타입]')
        type_cell.font = desc_header_font
        type_cell.fill = desc_header_fill
        type_cell.border = thin_border
        desc_cell = ws.cell(row=desc_start, column=3, value='[설명]')
        desc_cell.font = desc_header_font
        desc_cell.fill = desc_header_fill
        desc_cell.border = thin_border
        # Each column description
        for i, (col_name, col_type, col_desc) in enumerate(descriptions):
            r = desc_start + 1 + i
            c1 = ws.cell(row=r, column=1, value=col_name)
            c1.font = Font(bold=True, color='37474F', size=10)
            c1.fill = desc_fill
            c1.border = thin_border
            c2 = ws.cell(row=r, column=2, value=col_type)
            c2.font = desc_font
            c2.fill = desc_fill
            c2.border = thin_border
            c3 = ws.cell(row=r, column=3, value=col_desc)
            c3.font = desc_font
            c3.fill = desc_fill
            c3.border = thin_border
            c3.alignment = Alignment(wrap_text=True)
    # Auto column width (consider description text too)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


# ============================================================
# 1. TB_Currency
# ============================================================
ws = wb.active
ws.title = 'TB_Currency'
add_data(ws,
    ['currency_id', 'currency_name', 'currency_desc', 'daily_cap', 'icon_key'],
    [
        ['CUR_001', '골드', '차량 언락, 스탯 강화, 일부 커스텀', 800, 'ico_gold'],
        ['CUR_002', '스크랩', '파츠 합성 재료', 120, 'ico_scrap'],
    ],
    ['pk', '', '', '', ''],
    descriptions=[
        ('currency_id', 'string (PK)', '재화 고유 ID. 형식: CUR_###'),
        ('currency_name', 'string', '재화 표시 이름'),
        ('currency_desc', 'string', '재화 용도 설명 (UI 툴팁용)'),
        ('daily_cap', 'int', '하루 최대 획득 가능량. 0 = 무제한'),
        ('icon_key', 'string', 'UI 아이콘 리소스 키'),
    ]
)

# ============================================================
# 2. TB_Car
# ============================================================
ws = wb.create_sheet('TB_Car')
add_data(ws,
    ['car_id', 'car_name', 'car_type', 'base_hp', 'base_speed', 'base_atk_speed',
     'base_damage', 'collision_damage', 'passive_type', 'passive_value',
     'passive_desc', 'unlock_cost', 'unlock_currency_id', 'unlocked_by_default', 'sprite_key'],
    [
        ['CAR_001', '스포츠카', 'SportsCar', 100, 5, 1.5, 15, 10,
         'CooldownReduce', 15, '공격 쿨타임 -15%', 0, 'CUR_001', True, 'spr_car_sports'],
        ['CAR_002', 'SUV', 'SUV', 140, 3.5, 1.0, 15, 12,
         'DefenseUp', 10, '기본 방어력 +10%', 500, 'CUR_001', False, 'spr_car_suv'],
        ['CAR_003', '트럭', 'Truck', 200, 2.5, 0.8, 15, 20,
         'CollisionReflect', 100, '충돌 반사 데미지', 1500, 'CUR_001', False, 'spr_car_truck'],
    ],
    ['pk', '', 'enum', '', '', '', '', '', 'enum', '', '', '', 'fk', '', ''],
    descriptions=[
        ('car_id', 'string (PK)', '차량 고유 ID. 형식: CAR_###'),
        ('car_name', 'string', '차량 표시 이름'),
        ('car_type', 'enum', '차량 타입. 값: SportsCar / SUV / Truck'),
        ('base_hp', 'float', '기본 체력 (HP). 파츠에 의해 증감'),
        ('base_speed', 'float', '기본 이동속도 (units/sec)'),
        ('base_atk_speed', 'float', '기본 공격속도 (shots/sec). 높을수록 빠름'),
        ('base_damage', 'float', '기본 투사체 1회 데미지'),
        ('collision_damage', 'float', '적과 충돌 시 적에게 가하는 데미지'),
        ('passive_type', 'enum', '고유 패시브 종류. 값: CooldownReduce / DefenseUp / CollisionReflect'),
        ('passive_value', 'float', '패시브 수치 (% 또는 고정값). passive_type에 따라 해석 다름'),
        ('passive_desc', 'string', '패시브 설명 (UI 표시용)'),
        ('unlock_cost', 'int', '해금 비용. 0 = 무료(기본 차량)'),
        ('unlock_currency_id', 'string (FK)', '해금에 사용하는 재화 ID. → TB_Currency.currency_id'),
        ('unlocked_by_default', 'bool', 'true = 처음부터 사용 가능'),
        ('sprite_key', 'string', '차량 스프라이트 리소스 키'),
    ]
)

# ============================================================
# 3. TB_PartGrade
# ============================================================
ws = wb.create_sheet('TB_PartGrade')
add_data(ws,
    ['grade_id', 'grade_name', 'grade_enum', 'value_multiplier', 'color_hex', 'sort_order'],
    [
        ['GRADE_001', '일반', 'Common', 1.0, '#FFFFFF', 1],
        ['GRADE_002', '레어', 'Rare', 1.3, '#4FC3F7', 2],
        ['GRADE_003', '에픽', 'Epic', 1.6, '#CE93D8', 3],
        ['GRADE_004', '레전더리', 'Legendary', 2.0, '#FFD54F', 4],
        ['GRADE_005', '진화', 'Evolution', 2.5, '#FF8A65', 5],
    ],
    ['pk', '', 'enum', '', '', ''],
    descriptions=[
        ('grade_id', 'string (PK)', '등급 고유 ID. 형식: GRADE_###'),
        ('grade_name', 'string', '등급 한글 표시 이름'),
        ('grade_enum', 'enum', '코드에서 사용하는 등급 enum. 값: Common / Rare / Epic / Legendary / Evolution'),
        ('value_multiplier', 'float', 'TB_Part.base_value에 곱하는 배율. 실제효과 = base_value * value_multiplier'),
        ('color_hex', 'string', 'UI에서 등급 표시 색상 (HEX)'),
        ('sort_order', 'int', '정렬 순서. 낮을수록 앞에 표시'),
    ]
)

# ============================================================
# 4. TB_Part (15 base + 5 evolution = 20)
# ============================================================
ws = wb.create_sheet('TB_Part')
add_data(ws,
    ['part_id', 'part_name', 'category', 'base_value', 'effect_type', 'effect_desc',
     'has_active', 'weapon_type', 'cooldown', 'duration', 'max_level',
     'stackable', 'drop_weight', 'is_evolution_result', 'icon_key'],
    [
        # Engine (4)
        ['PART_001', '터보차저', 'Engine', 12, 'SpeedUp', '이동 속도 +%',
         False, 'None', 0, 0, 3, True, 10, False, 'ico_turbocharger'],
        ['PART_002', '슈퍼차저', 'Engine', 15, 'AtkSpeedUp', '공격 속도 +%',
         False, 'None', 0, 0, 3, True, 10, False, 'ico_supercharger'],
        ['PART_003', 'NOS 부스터', 'Engine', 20, 'ActiveSkill', '3초 무적 돌진 (쿨타임 초)',
         True, 'None', 20, 3, 3, True, 8, False, 'ico_nos'],
        ['PART_004', '경량화 섀시', 'Engine', 8, 'SpeedUp', '이동 속도 +%, 체력 -5%',
         False, 'None', 0, 0, 3, True, 8, False, 'ico_lightweight'],
        # Weapon (5)
        ['PART_005', '기관총', 'Weapon', 20, 'DamageUp', '초당 데미지',
         False, 'MachineGun', 0, 0, 3, True, 10, False, 'ico_machinegun'],
        ['PART_006', '미사일 런처', 'Weapon', 60, 'DamageUp', '발사체 데미지',
         False, 'Missile', 3, 0, 3, True, 8, False, 'ico_missile'],
        ['PART_007', 'EMP 펄스', 'Weapon', 3, 'ActiveSkill', '마비 범위 (units)',
         True, 'EMP', 8, 1, 3, True, 7, False, 'ico_emp'],
        ['PART_008', '오일 슬릭', 'Weapon', 2, 'ActiveSkill', '슬립 지속 (초)',
         True, 'OilSlick', 5, 5, 3, True, 7, False, 'ico_oilslick'],
        ['PART_009', '지뢰 투하', 'Weapon', 80, 'DamageUp', '지뢰 데미지',
         False, 'Mine', 4, 10, 3, True, 7, False, 'ico_mine'],
        # Defense (3)
        ['PART_010', '런플랫 타이어', 'Defense', 20, 'DefenseUp', '피격 데미지 감소 %',
         False, 'None', 0, 0, 3, True, 10, False, 'ico_runflat'],
        ['PART_011', '강화 차체', 'Defense', 30, 'HpUp', '최대 체력 +%',
         False, 'None', 0, 0, 3, True, 10, False, 'ico_reinforced'],
        ['PART_012', '사이드 램', 'Defense', 50, 'CollisionReflect', '충돌 반사 데미지 %',
         False, 'None', 0, 0, 3, True, 7, False, 'ico_sideram'],
        # Special (3)
        ['PART_013', '드래프팅', 'Special', 25, 'SpeedUp', '후방 접근 속도 +%',
         False, 'None', 0, 0, 3, True, 7, False, 'ico_drafting'],
        ['PART_014', '자동 수리 킷', 'Special', 1, 'HpRegen', '초당 체력 회복 %',
         False, 'None', 0, 0, 3, True, 6, False, 'ico_autorepair'],
        ['PART_015', '레이더', 'Special', 20, 'ActiveSkill', '미니맵 탐지 범위 (units)',
         True, 'None', 0, 0, 3, True, 4, False, 'ico_radar'],
        # Evolution Results (5)
        ['PART_016', '트윈터보', 'Engine', 20, 'SpeedUp', '속도+공격속도 동시 +20%',
         False, 'None', 0, 0, 5, False, 0, True, 'ico_twintorbo'],
        ['PART_017', '유도 EMP 미사일', 'Weapon', 90, 'DamageUp', '유도 + 광역 마비',
         False, 'Missile', 3, 1, 5, False, 0, True, 'ico_guided_emp'],
        ['PART_018', '화염 기관총', 'Weapon', 30, 'DamageUp', '이동 경로에 화염 지속 데미지',
         False, 'MachineGun', 0, 5, 5, False, 0, True, 'ico_flame_mg'],
        ['PART_019', '불도저 모드', 'Defense', 60, 'CollisionReflect', '충돌 시 무적 + 반사 데미지',
         True, 'None', 15, 3, 5, False, 0, True, 'ico_bulldozer'],
        ['PART_020', '슬립스트림 폭발', 'Special', 40, 'ActiveSkill', '돌진 시 연쇄 폭발 발생',
         True, 'None', 18, 3, 5, False, 0, True, 'ico_slipstream'],
    ],
    ['pk', '', 'enum', '', 'enum', '', '', 'enum', '', '', '', '', '', '', ''],
    descriptions=[
        ('part_id', 'string (PK)', '파츠 고유 ID. 형식: PART_###'),
        ('part_name', 'string', '파츠 표시 이름'),
        ('category', 'enum', '파츠 카테고리. 값: Engine / Weapon / Defense / Special'),
        ('base_value', 'float', '일반(Common) 등급 기준 효과 수치. 등급별 실제값 = base_value * TB_PartGrade.value_multiplier'),
        ('effect_type', 'enum', '효과 종류. 값: SpeedUp / AtkSpeedUp / DamageUp / HpUp / DefenseUp / HpRegen / CollisionReflect / ActiveSkill'),
        ('effect_desc', 'string', '효과 설명 (UI 카드 표시용)'),
        ('has_active', 'bool', 'true = 액티브 스킬 보유. cooldown/duration 필드 사용'),
        ('weapon_type', 'enum', '무기 카테고리 파츠의 발사 타입. 값: None / MachineGun / Missile / EMP / OilSlick / Mine'),
        ('cooldown', 'float', '액티브 스킬 재사용 대기시간 (초). has_active=false면 0'),
        ('duration', 'float', '액티브 스킬 지속시간 (초). has_active=false면 0'),
        ('max_level', 'int', '파츠 최대 레벨. 일반=3, 진화=5'),
        ('stackable', 'bool', 'true = 같은 파츠 중복 장착 가능'),
        ('drop_weight', 'int', '레벨업 시 등장 가중치. 높을수록 자주 등장. 진화 파츠=0 (드랍 불가)'),
        ('is_evolution_result', 'bool', 'true = 진화로만 획득 가능한 파츠. 드랍/상점 불가'),
        ('icon_key', 'string', '파츠 아이콘 리소스 키'),
    ]
)

# ============================================================
# 5. TB_Evolution
# ============================================================
ws = wb.create_sheet('TB_Evolution')
add_data(ws,
    ['evo_id', 'evo_name', 'material_a_id', 'material_a_level',
     'material_b_id', 'material_b_level', 'result_part_id', 'visual_desc', 'priority'],
    [
        ['EVO_001', '트윈터보', 'PART_001', 3, 'PART_002', 1, 'PART_016',
         '지붕에 쌍 터보 스쿠프 추가', 'P0'],
        ['EVO_002', '유도 EMP 미사일', 'PART_006', 3, 'PART_007', 1, 'PART_017',
         '차 양옆 미사일 포드 장착', 'P0'],
        ['EVO_003', '화염 기관총', 'PART_008', 2, 'PART_005', 2, 'PART_018',
         '차 전면 화염방사기 장착', 'P0'],
        ['EVO_004', '불도저 모드', 'PART_011', 3, 'PART_012', 1, 'PART_019',
         '차 전면 거대 불도저 날 추가', 'P1'],
        ['EVO_005', '슬립스트림 폭발', 'PART_003', 1, 'PART_013', 1, 'PART_020',
         '차 후미 NOS 탱크 + 불꽃 노즐', 'P1'],
    ],
    ['pk', '', 'fk', '', 'fk', '', 'fk', '', 'enum'],
    descriptions=[
        ('evo_id', 'string (PK)', '진화 레시피 고유 ID. 형식: EVO_###'),
        ('evo_name', 'string', '진화 결과 파츠 이름 (표시용)'),
        ('material_a_id', 'string (FK)', '재료 A 파츠 ID. → TB_Part.part_id'),
        ('material_a_level', 'int', '재료 A가 도달해야 하는 최소 레벨'),
        ('material_b_id', 'string (FK)', '재료 B 파츠 ID. → TB_Part.part_id'),
        ('material_b_level', 'int', '재료 B가 도달해야 하는 최소 레벨'),
        ('result_part_id', 'string (FK)', '진화 결과 파츠 ID. → TB_Part.part_id (is_evolution_result=true인 파츠)'),
        ('visual_desc', 'string', '진화 시 차량 외형 변화 설명. 아트팀 전달용'),
        ('priority', 'enum', '개발 우선순위. 값: P0 (MVP 필수) / P1 (중요) / P2 (후순위)'),
    ]
)

# ============================================================
# 6. TB_Monster
# ============================================================
ws = wb.create_sheet('TB_Monster')
add_data(ws,
    ['mon_id', 'mon_name', 'is_boss', 'base_hp', 'base_speed', 'contact_damage',
     'scale', 'spawn_start_min', 'spawn_weight', 'special_ability', 'tint_color', 'sprite_key'],
    [
        ['MON_001', '러스티 세단', False, 60, 2, 8, 1.0, 0, 10, 'None', '#8B4513', 'spr_mon_rusty'],
        ['MON_002', '스파이크 바이크', False, 30, 5, 12, 0.8, 2, 7, 'None', '#C0C0C0', 'spr_mon_spike'],
        ['MON_003', '장갑 픽업', False, 150, 1.5, 15, 1.2, 4, 5, 'None', '#556B2F', 'spr_mon_armored'],
        ['MON_004', '불꽃 버기', False, 50, 4.5, 10, 0.9, 7, 4, 'FlameTrail', '#FF4500', 'spr_mon_flame'],
        ['MON_005', '워로드 탱크트럭', True, 2000, 1, 30, 2.0, 12, 1, 'ThreePhase', '#8B0000', 'spr_mon_warlord'],
    ],
    ['pk', '', '', '', '', '', '', '', '', '', '', ''],
    descriptions=[
        ('mon_id', 'string (PK)', '적 고유 ID. 형식: MON_###'),
        ('mon_name', 'string', '적 표시 이름'),
        ('is_boss', 'bool', 'true = 보스 몬스터. 보스는 TB_Wave에서 단독 등장'),
        ('base_hp', 'float', '기본 체력. TB_Wave.difficulty_scale에 의해 스케일링됨'),
        ('base_speed', 'float', '이동속도 (units/sec). 난이도 배율 적용됨'),
        ('contact_damage', 'float', '플레이어와 충돌 시 프레임당 데미지'),
        ('scale', 'float', '스프라이트 크기 배율. 1.0 = 기본 크기'),
        ('spawn_start_min', 'int', '이 적이 처음 등장하는 시간 (분). 0 = 게임 시작부터'),
        ('spawn_weight', 'int', '스폰 가중치. 높을수록 해당 시간대에 자주 스폰'),
        ('special_ability', 'string', '특수 능력. None = 없음 / FlameTrail = 화염 궤적 / ThreePhase = 3페이즈 보스'),
        ('tint_color', 'string', '스프라이트 색조 (HEX). 시각적 구분용'),
        ('sprite_key', 'string', '스프라이트 리소스 키'),
    ]
)

# ============================================================
# 7. TB_MonsterDrop
# ============================================================
ws = wb.create_sheet('TB_MonsterDrop')
add_data(ws,
    ['drop_id', 'mon_id', 'exp_amount', 'gold_amount', 'special_drop_id', 'special_drop_rate'],
    [
        ['DROP_001', 'MON_001', 3, 5, None, 0],
        ['DROP_002', 'MON_002', 2, 3, None, 0],
        ['DROP_003', 'MON_003', 5, 8, None, 0],
        ['DROP_004', 'MON_004', 4, 6, None, 0],
        ['DROP_005', 'MON_005', 50, 200, None, 0],
    ],
    ['pk', 'fk', '', '', 'fk', ''],
    descriptions=[
        ('drop_id', 'string (PK)', '드랍 고유 ID. 형식: DROP_###'),
        ('mon_id', 'string (FK)', '이 드랍을 가진 적 ID. → TB_Monster.mon_id'),
        ('exp_amount', 'int', '처치 시 드랍하는 경험치(파편) 수량'),
        ('gold_amount', 'int', '처치 시 드랍하는 골드 수량'),
        ('special_drop_id', 'string (FK, nullable)', '특수 드랍 파츠 ID. → TB_Part.part_id. 없으면 null'),
        ('special_drop_rate', 'float', '특수 드랍 확률 (%). 0 = 특수 드랍 없음'),
    ]
)

# ============================================================
# 8. TB_Wave (special handling for multi-monster rows)
# ============================================================
ws = wb.create_sheet('TB_Wave')
wave_headers = ['wave_id', 'time_min', 'phase_name', 'mon_id', 'spawn_per_30s',
                'max_enemies', 'difficulty_scale', 'note']
wave_col_types = ['pk', '', '', 'fk', '', '', '', '']
wave_descriptions = [
    ('wave_id', 'string (PK)', '웨이브 고유 ID. 형식: WAVE_###. 같은 시간대라도 적 종류별로 행이 나뉨'),
    ('time_min', 'int', '웨이브 시작 시간 (분). 0 = 게임 시작'),
    ('phase_name', 'string', '구간 이름 (UI/디버그용). 튜토리얼/초반/중반/후반/위기/보스전/런종료'),
    ('mon_id', 'string (FK)', '이 웨이브에서 스폰하는 적 ID. → TB_Monster.mon_id'),
    ('spawn_per_30s', 'int', '30초당 스폰 수. 0 = 추가 스폰 없음 (보스전 등)'),
    ('max_enemies', 'int', '해당 시간대 동시 존재 가능한 최대 적 수'),
    ('difficulty_scale', 'float', '난이도 배율. 적 HP/속도에 곱해짐. 1.0 = 기본'),
    ('note', 'string', '특이사항 메모'),
]

# Write header
ws.append(wave_headers)
for col_idx, h in enumerate(wave_headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

wave_timeline = [
    (0, '튜토리얼', [('MON_001', 4)], 15, 1.0, '적응 구간, 러스티 세단만'),
    (1, '초반', [('MON_001', 5)], 20, 1.1, ''),
    (2, '초반', [('MON_001', 5), ('MON_002', 2)], 25, 1.2, '스파이크 바이크 첫 등장'),
    (3, '중반 시작', [('MON_001', 4), ('MON_002', 3), ('MON_003', 1)], 30, 1.3, '장갑 픽업 첫 등장'),
    (4, '중반', [('MON_001', 4), ('MON_002', 3), ('MON_003', 1)], 35, 1.4, ''),
    (5, '중반', [('MON_001', 5), ('MON_002', 4), ('MON_003', 2)], 40, 1.5, '밀도 증가'),
    (6, '중반 후반', [('MON_001', 5), ('MON_002', 4), ('MON_003', 2)], 45, 1.6, ''),
    (7, '후반 시작', [('MON_001', 4), ('MON_002', 3), ('MON_003', 2), ('MON_004', 2)], 50, 1.8, '불꽃 버기 첫 등장'),
    (8, '후반', [('MON_001', 4), ('MON_002', 4), ('MON_003', 2), ('MON_004', 2)], 55, 1.9, ''),
    (9, '후반', [('MON_001', 5), ('MON_002', 4), ('MON_003', 3), ('MON_004', 3)], 60, 2.0, ''),
    (10, '위기', [('MON_001', 5), ('MON_002', 5), ('MON_003', 3), ('MON_004', 3)], 65, 2.2, ''),
    (11, '위기', [('MON_001', 6), ('MON_002', 5), ('MON_003', 3), ('MON_004', 4)], 70, 2.3, ''),
    (12, '보스 등장!', [('MON_005', 1)], 1, 3.0, '워로드 탱크트럭 등장 - 기존 적 전부 사라짐'),
    (13, '보스전', [('MON_005', 0)], 1, 3.0, '보스 처치 시 런 완료'),
    (14, '보스전', [('MON_005', 0)], 1, 3.0, '미처치 시 연료 게이지 마저 소진'),
    (15, '런 종료', [], 0, 0, '연료 소진 - 강제 런 종료'),
]

wave_idx = 0
data_count = 0
for time_min, phase, monsters, max_e, diff, note in wave_timeline:
    if not monsters:
        ws.append(["WAVE_{:03d}".format(wave_idx), time_min, phase, '', 0, max_e, diff, note])
        wave_idx += 1
        data_count += 1
    else:
        for mon_id, spawn in monsters:
            ws.append(["WAVE_{:03d}".format(wave_idx), time_min, phase, mon_id, spawn, max_e, diff, note])
            wave_idx += 1
            data_count += 1

# Style wave data rows
for row in ws.iter_rows(min_row=2, max_row=1 + data_count, max_col=len(wave_headers)):
    for col_idx, cell in enumerate(row, 1):
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx - 1 < len(wave_col_types):
            t = wave_col_types[col_idx - 1]
            if t == 'pk':
                cell.fill = pk_fill
            elif t == 'fk':
                cell.fill = fk_fill

# Wave descriptions
desc_start = 1 + data_count + 2
label_cell = ws.cell(row=desc_start, column=1, value='[컬럼명]')
label_cell.font = desc_header_font
label_cell.fill = desc_header_fill
label_cell.border = thin_border
type_cell = ws.cell(row=desc_start, column=2, value='[타입]')
type_cell.font = desc_header_font
type_cell.fill = desc_header_fill
type_cell.border = thin_border
dc = ws.cell(row=desc_start, column=3, value='[설명]')
dc.font = desc_header_font
dc.fill = desc_header_fill
dc.border = thin_border
for i, (col_name, col_type, col_desc) in enumerate(wave_descriptions):
    r = desc_start + 1 + i
    c1 = ws.cell(row=r, column=1, value=col_name)
    c1.font = Font(bold=True, color='37474F', size=10)
    c1.fill = desc_fill
    c1.border = thin_border
    c2 = ws.cell(row=r, column=2, value=col_type)
    c2.font = desc_font
    c2.fill = desc_fill
    c2.border = thin_border
    c3 = ws.cell(row=r, column=3, value=col_desc)
    c3.font = desc_font
    c3.fill = desc_fill
    c3.border = thin_border
    c3.alignment = Alignment(wrap_text=True)

for col_idx in range(1, len(wave_headers) + 1):
    max_len = len(str(wave_headers[col_idx - 1]))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

# ============================================================
# 9. TB_Level
# ============================================================
ws = wb.create_sheet('TB_Level')
add_data(ws,
    ['level', 'required_exp', 'required_exp_gap', 'difficulty_multiplier',
     'grade_common_pct', 'grade_rare_pct', 'grade_epic_pct', 'grade_legendary_pct', 'note'],
    [
        [1, 30, 30, 1.0, 70, 25, 5, 0, '초반 튜토리얼 구간'],
        [2, 70, 40, 1.1, 65, 28, 7, 0, ''],
        [3, 120, 50, 1.2, 50, 35, 13, 2, '일반/레어 혼합'],
        [4, 180, 60, 1.3, 30, 45, 20, 5, '레어 위주'],
        [5, 250, 70, 1.5, 25, 40, 25, 10, '5레벨마다 진화 재료 확률 UP'],
        [6, 330, 80, 1.6, 20, 35, 30, 15, '레어/에픽 혼합'],
        [7, 420, 90, 1.8, 15, 30, 35, 20, '레어/에픽 혼합'],
        [8, 520, 100, 2.0, 10, 20, 45, 25, '에픽 위주'],
        [9, 630, 110, 2.2, 5, 15, 45, 35, '에픽 위주'],
        [10, 750, 120, 2.5, 0, 10, 40, 50, '보스 소환 트리거'],
    ],
    ['pk', '', '', '', '', '', '', '', ''],
    descriptions=[
        ('level', 'int (PK)', '플레이어 레벨. 1부터 시작'),
        ('required_exp', 'int', '해당 레벨 도달에 필요한 누적 경험치(파편)'),
        ('required_exp_gap', 'int', '이전 레벨 대비 추가로 필요한 경험치 (구간값)'),
        ('difficulty_multiplier', 'float', '해당 레벨 시점의 웨이브 난이도 배율'),
        ('grade_common_pct', 'int', '레벨업 시 일반 등급 파츠 등장 확률 (%)'),
        ('grade_rare_pct', 'int', '레벨업 시 레어 등급 파츠 등장 확률 (%)'),
        ('grade_epic_pct', 'int', '레벨업 시 에픽 등급 파츠 등장 확률 (%)'),
        ('grade_legendary_pct', 'int', '레벨업 시 레전더리 등급 파츠 등장 확률 (%). 4개 합계 = 100'),
        ('note', 'string', '비고'),
    ]
)

# ============================================================
# 10. TB_Reward
# ============================================================
ws = wb.create_sheet('TB_Reward')
add_data(ws,
    ['reward_id', 'reward_trigger', 'currency_id', 'amount', 'bonus_multiplier', 'note'],
    [
        ['RWD_001', 'RunComplete', 'CUR_001', 100, 1.0, '런 완료 기본 골드'],
        ['RWD_002', 'RunComplete', 'CUR_002', 20, 1.0, '런 완료 기본 스크랩'],
        ['RWD_003', 'BossClear', 'CUR_001', 200, 1.0, '보스 처치 보너스 골드'],
        ['RWD_004', 'BossClear', 'CUR_002', 50, 1.0, '보스 처치 보너스 스크랩'],
        ['RWD_005', 'AdWatch', 'CUR_001', 200, 2.0, '광고 시청 시 골드 2배'],
        ['RWD_006', 'DailyLogin', 'CUR_001', 50, 1.0, '일일 접속 골드'],
        ['RWD_007', 'DailyLogin', 'CUR_002', 10, 1.0, '일일 접속 스크랩'],
    ],
    ['pk', 'enum', 'fk', '', '', ''],
    descriptions=[
        ('reward_id', 'string (PK)', '보상 고유 ID. 형식: RWD_###'),
        ('reward_trigger', 'enum', '보상 발동 조건. 값: RunComplete / BossClear / AdWatch / DailyLogin'),
        ('currency_id', 'string (FK)', '지급할 재화 ID. → TB_Currency.currency_id'),
        ('amount', 'int', '기본 지급 수량'),
        ('bonus_multiplier', 'float', '보너스 배율. 1.0 = 기본. 광고 시청 등으로 2.0 적용 가능'),
        ('note', 'string', '비고'),
    ]
)

# ============================================================
# 11. TB_Shop
# ============================================================
ws = wb.create_sheet('TB_Shop')
add_data(ws,
    ['shop_id', 'shop_name', 'shop_type', 'price_type', 'price_amount',
     'price_currency_id', 'reward_type', 'reward_target_id', 'reward_amount',
     'is_one_time', 'release_phase', 'note'],
    [
        ['SHOP_001', '스타터 패키지', 'IAP', 'RealMoney', 1900, None,
         'Currency', 'CUR_001', 1000, True, 'MVP', '첫 구매 유도용 (+ 레어 파츠 3개)'],
        ['SHOP_002', '프리미엄 도색 팩', 'IAP', 'RealMoney', 2900, None,
         'Cosmetic', 'COSMETIC_PAINT', 10, True, 'MVP', '메탈릭/크롬/무광 색상 10종'],
        ['SHOP_003', '프리미엄 휠 팩', 'IAP', 'RealMoney', 2900, None,
         'Cosmetic', 'COSMETIC_WHEEL', 5, True, 'MVP', '특수 휠 디자인 5종'],
        ['SHOP_004', '광고 제거', 'IAP', 'RealMoney', 4900, None,
         'System', 'SYS_ADREMOVE', 1, True, 'MVP', '전면 광고 완전 제거'],
        ['SHOP_005', '월정액 카드', 'IAP', 'RealMoney', 4900, None,
         'Currency', 'CUR_001', 200, False, 'v1.5', '매일 골드 200 + 스크랩 50'],
        ['SHOP_006', '시즌 번들', 'IAP', 'RealMoney', 9900, None,
         'Cosmetic', 'COSMETIC_SEASON', 1, True, 'v2', '한정 도색 + 휠 + 바디킷 세트'],
        ['SHOP_007', 'SUV 언락', 'InGame', 'Gold', 500, 'CUR_001',
         'Car', 'CAR_002', 1, True, 'MVP', '인게임 골드로 SUV 구매'],
        ['SHOP_008', '트럭 언락', 'InGame', 'Gold', 1500, 'CUR_001',
         'Car', 'CAR_003', 1, True, 'MVP', '인게임 골드로 트럭 구매'],
    ],
    ['pk', '', 'enum', 'enum', '', 'fk', 'enum', 'fk', '', '', 'enum', ''],
    descriptions=[
        ('shop_id', 'string (PK)', '상품 고유 ID. 형식: SHOP_###'),
        ('shop_name', 'string', '상품 표시 이름'),
        ('shop_type', 'enum', '상품 종류. 값: IAP (실제 결제) / InGame (인게임 재화) / Ad (광고 보상)'),
        ('price_type', 'enum', '결제 수단. 값: RealMoney (원화) / Gold / Scrap'),
        ('price_amount', 'int', '가격. RealMoney일 때 원(KRW) 단위'),
        ('price_currency_id', 'string (FK, nullable)', '인게임 재화 결제 시 재화 ID. → TB_Currency.currency_id. IAP면 null'),
        ('reward_type', 'enum', '보상 종류. 값: Currency / Part / Car / Cosmetic / System'),
        ('reward_target_id', 'string (FK)', '보상 대상 ID. reward_type에 따라 참조 테이블이 다름 (TB_Currency / TB_Part / TB_Car 등)'),
        ('reward_amount', 'int', '보상 수량'),
        ('is_one_time', 'bool', 'true = 1회 구매 제한. false = 반복 구매 가능'),
        ('release_phase', 'enum', '출시 단계. 값: MVP / v1.5 / v2'),
        ('note', 'string', '비고'),
    ]
)

# ============================================================
# Save
# ============================================================
wb.save('car_survivor_tables.xlsx')
print('car_survivor_tables.xlsx created!')
print("Total sheets: {}".format(len(wb.sheetnames)))
for name in wb.sheetnames:
    sheet = wb[name]
    print("  {}: {} rows".format(name, sheet.max_row - 1))
