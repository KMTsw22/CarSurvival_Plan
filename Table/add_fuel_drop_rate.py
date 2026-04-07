"""
TB_MonsterDrop에 fuel_drop_rate 컬럼 추가 (0.1%) 후 재익스포트

사용법: py add_fuel_drop_rate.py
"""
import os, sys, io, msgpack, openpyxl
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

DROP_TYPES = [str, str, int, int, int, float]

def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_MonsterDrop"]

    # 헤더 추가
    col_idx = 6
    header_cell = ws.cell(row=1, column=col_idx)
    header_cell.value = "fuel_drop_rate"
    header_cell.comment = Comment("연료통 드롭 확률 (%)\n0.1 = 0.1%", "System")

    # 각 행에 0.1 설정
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value is None: break
        ws.cell(row=row_idx, column=col_idx, value=0.1)
        print(f"{row[0].value} ({row[1].value}): fuel_drop_rate = 0.1%")

    wb.save(XLSX_PATH)
    wb.close()
    print("\n엑셀 저장 완료!")

    # 재익스포트
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_MonsterDrop"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["): break
        rows.append([cast(row[i] if i < len(row) else None, DROP_TYPES[i]) for i in range(len(DROP_TYPES))])
    wb.close()

    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_MonsterDrop.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)
    print(f"TB_MonsterDrop.bytes: {len(rows)} rows, {len(packed)} bytes")

if __name__ == "__main__":
    main()
