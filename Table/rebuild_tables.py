"""
MVP 기준으로 엑셀 테이블 재구성 스크립트
삭제: TB_Shop, TB_Reward, TB_Evolution, TB_PartGrade
수정: TB_Currency, TB_Car, TB_Part, TB_Monster, TB_MonsterDrop, TB_Wave, TB_Level, TB_Map
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
desc_header_font = Font(bold=True, color="666666", size=10)
desc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def write_sheet(wb, name, headers, data, desc_rows=None):
    ws = wb.create_sheet(name)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, row_data in enumerate(data, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    if desc_rows:
        start = len(data) + 3
        for c_idx, dh in enumerate(["[컬럼명]", "[타입]", "[설명]"], 1):
            cell = ws.cell(row=start, column=c_idx, value=dh)
            cell.font = desc_header_font
            cell.fill = desc_fill
        for r_idx, desc in enumerate(desc_rows, start + 1):
            for c_idx, val in enumerate(desc, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


wb = openpyxl.Workbook()
wb.remove(wb.active)

# === TB_Currency ===
write_sheet(wb, "TB_Currency",
    ["currency_id", "currency_name", "currency_desc", "icon_key"],
    [
        ["CUR_001", "골드", "업그레이드 비용, 상점 구매", "ico_gold"],
        ["CUR_002", "나사", "장비 업그레이드 재료", "ico_screw"],
    ],
    [
        ["currency_id", "string (PK)", "재화 고유 ID. 규칙: CUR_###"],
        ["currency_name", "string", "재화 표시 이름"],
        ["currency_desc", "string", "재화 용도 설명"],
        ["icon_key", "string", "UI 아이콘 리소스 키"],
    ]
)

# === TB_Car ===
write_sheet(wb, "TB_Car",
    ["car_id", "car_name", "car_type", "base_hp", "base_speed", "base_atk_speed",
     "base_damage", "collision_damage", "passive_type", "passive_value", "passive_desc",
     "unlock_cost", "unlock_currency_id", "unlocked_by_default", "sprite_key"],
    [
        ["CAR_001", "슈퍼카", "SuperCar", 100, 6.0, 1.5, 15, 8,
         "CooldownReduce", 15, "스킬 쿨타임 -15%", 0, "CUR_001", True, "spr_car_super"],
        ["CAR_002", "SUV", "SUV", 150, 4.0, 1.0, 15, 15,
         "DefenseUp", 10, "기본 방어 +10%", 500, "CUR_001", False, "spr_car_suv"],
        ["CAR_003", "오토바이", "Motorcycle", 80, 7.0, 1.8, 12, 5,
         "SpeedUp", 20, "이동속도 +20%", 1000, "CUR_001", False, "spr_car_motorcycle"],
    ],
    [
        ["car_id", "string (PK)", "차량 고유 ID"],
        ["car_name", "string", "차량 표시 이름"],
        ["car_type", "enum", "차량 타입. SuperCar / SUV / Motorcycle"],
        ["base_hp", "float", "기본 체력"],
        ["base_speed", "float", "기본 이동속도 (units/sec)"],
        ["base_atk_speed", "float", "기본 공격속도 (shots/sec)"],
        ["base_damage", "float", "기본 투사체 데미지"],
        ["collision_damage", "float", "차량 충돌 데미지"],
        ["passive_type", "enum", "차량 패시브 종류"],
        ["passive_value", "float", "패시브 수치"],
        ["passive_desc", "string", "패시브 설명"],
        ["unlock_cost", "int", "해금 비용. 0=무료"],
        ["unlock_currency_id", "string (FK)", "해금 재화 ID"],
        ["unlocked_by_default", "bool", "true=기본 해금"],
        ["sprite_key", "string", "스프라이트 리소스 키"],
    ]
)

# === TB_Weapon (주무기 + 보조무기) ===
# etc_value 주석:
#   기관총: etc1=없음, etc2=없음, etc3=없음
#   오일슬릭: etc1=감속비율(%), etc2=감속지속(초), etc3=범위반경
#   회전톱날: etc1=회전속도(도/초), etc2=궤도반경, etc3=타격간격(초)
#   체인라이트닝: etc1=기본타격인원, etc2=레벨당추가인원, etc3=없음
#   EMP펄스: etc1=기본지속시간(초), etc2=기본반경, etc3=레벨당반경증가
#   화염방사기: etc1=기본지속시간(초), etc2=기본반경, etc3=레벨당반경증가
#   레이저캐논: etc1=기본레이저수, etc2=레벨당레이저추가, etc3=없음
#   미사일포드: etc1=기본미사일수, etc2=레벨당미사일추가, etc3=없음
write_sheet(wb, "TB_Weapon",
    ["weapon_id", "weapon_name", "weapon_category", "damage", "effect_desc",
     "aim_type", "weapon_type", "cooldown", "duration", "max_level",
     "drop_weight", "icon_key",
     "etc_value1", "etc_value2", "etc_value3", "etc_value4", "etc_value5",
     "damage_per_level"],
    [
        # 기관총: damage=기본뎀, damage_per_level=레벨당 추가 데미지
        ["WPN_001", "기관총", "Main", 20, "마우스 방향 연사",
         "Manual", "MachineGun", 0.3, 0, 5, 10, "ico_machinegun",
         0, 0, 0, 0, 0, 5],
        # 오일슬릭: damage=초당DPS, etc1=감속50%, etc2=감속0.5초, etc3=기본반경0.8
        ["WPN_002", "오일 슬릭", "Sub", 3, "독 웅덩이 (지속뎀+감속)",
         "Auto", "OilSlick", 5, 5, 5, 8, "ico_oilslick",
         50, 0.5, 0.8, 0, 0, 1],
        # 회전톱날: damage=타격당뎀, etc1=회전속도180, etc2=궤도반경2, etc3=타격간격0.3
        ["WPN_003", "회전 톱날", "Sub", 15, "차량 주변 회전 데미지",
         "Auto", "SawBlade", 0, 0, 5, 8, "ico_sawblade",
         180, 2, 0.3, 0, 0, 3],
        # 체인라이트닝: etc1=기본타격인원3, etc2=레벨당+1명
        ["WPN_004", "체인 라이트닝", "Sub", 25, "연쇄 번개 공격",
         "Auto", "ChainLightning", 1.5, 0, 5, 8, "ico_chainlightning",
         3, 1, 0, 0, 0, 5],
        # EMP펄스: etc1=기본지속시간2초, etc2=기본반경3, etc3=레벨당반경+0.5
        ["WPN_005", "EMP 펄스", "Sub", 10, "범위 전자기 충격",
         "Auto", "EMPPulse", 6, 2, 5, 8, "ico_emppulse",
         2, 3, 0.5, 0, 0, 3],
        # 화염방사기: etc1=기본지속시간3초, etc2=기본반경2, etc3=레벨당반경+0.3
        ["WPN_006", "화염방사기", "Sub", 8, "전방 화염 방사",
         "Auto", "Flamethrower", 4, 3, 5, 8, "ico_flamethrower",
         3, 2, 0.3, 0, 0, 2],
        # 레이저캐논: etc1=기본레이저1개, etc2=레벨당+1개
        ["WPN_007", "레이저 캐논", "Main", 30, "관통 레이저 발사",
         "Manual", "LaserCannon", 2, 0, 5, 8, "ico_lasercannon",
         1, 1, 0, 0, 0, 8],
        # 미사일포드: etc1=기본미사일2개, etc2=레벨당+1개
        ["WPN_008", "미사일 포드", "Sub", 40, "유도 미사일 발사",
         "Auto", "MissilePod", 3, 0, 5, 8, "ico_missilepod",
         2, 1, 0, 0, 0, 10],
    ],
    [
        ["weapon_id", "string (PK)", "무기 고유 ID. 규칙: WPN_###"],
        ["weapon_name", "string", "무기 표시 이름"],
        ["weapon_category", "enum", "Main=주무기 / Sub=보조무기"],
        ["damage", "float", "기본 데미지 (레벨 1 기준)"],
        ["effect_desc", "string", "효과 설명"],
        ["aim_type", "enum", "Manual=마우스 조준 / Auto=자동 발동"],
        ["weapon_type", "enum", "무기 타입. MachineGun/OilSlick/SawBlade/ChainLightning/EMPPulse/Flamethrower/LaserCannon/MissilePod"],
        ["cooldown", "float", "쿨타임/발사간격 (초). 0=없음"],
        ["duration", "float", "지속시간 (초). 0=영구"],
        ["max_level", "int", "최대 레벨"],
        ["drop_weight", "int", "레벨업 선택지 등장 가중치"],
        ["icon_key", "string", "아이콘 리소스 키"],
        ["etc_value1", "float", "무기별 추가값1 (주석 참조)"],
        ["etc_value2", "float", "무기별 추가값2 (주석 참조)"],
        ["etc_value3", "float", "무기별 추가값3 (주석 참조)"],
        ["etc_value4", "float", "무기별 추가값4 (예비)"],
        ["etc_value5", "float", "무기별 추가값5 (예비)"],
        ["damage_per_level", "float", "레벨당 데미지 증가량"],
    ]
)

# === TB_SpellBook (마법서 — 패시브 버프) ===
write_sheet(wb, "TB_SpellBook",
    ["book_id", "book_name", "effect_type", "base_value", "effect_desc",
     "max_level", "stackable", "drop_weight", "icon_key"],
    [
        ["BOOK_001", "이동속도 증가", "SpeedUp", 10, "이동속도 +%",
         5, True, 10, "ico_speed"],
        ["BOOK_002", "공격력 증가", "DamageUp", 10, "공격력 +%",
         5, True, 10, "ico_attack"],
        ["BOOK_003", "최대 체력 증가", "HpUp", 10, "최대 체력 +%",
         5, True, 10, "ico_hp"],
    ],
    [
        ["book_id", "string (PK)", "마법서 고유 ID. 규칙: BOOK_###"],
        ["book_name", "string", "마법서 표시 이름"],
        ["effect_type", "enum", "효과 타입. SpeedUp / DamageUp / HpUp"],
        ["base_value", "float", "기본 수치 (레벨당 증가 기준)"],
        ["effect_desc", "string", "효과 설명"],
        ["max_level", "int", "최대 레벨"],
        ["stackable", "bool", "중복 획득 가능 여부"],
        ["drop_weight", "int", "레벨업 선택지 등장 가중치"],
        ["icon_key", "string", "아이콘 리소스 키"],
    ]
)

# === TB_Part (빈 테이블 — 향후 사용) ===
write_sheet(wb, "TB_Part",
    ["part_id", "part_name", "part_desc", "effect_type", "effect_value",
     "max_level", "icon_key"],
    [],
    [
        ["part_id", "string (PK)", "파츠 고유 ID. 규칙: PART_###"],
        ["part_name", "string", "파츠 표시 이름"],
        ["part_desc", "string", "파츠 설명"],
        ["effect_type", "enum", "효과 타입 (미정)"],
        ["effect_value", "float", "효과 수치"],
        ["max_level", "int", "최대 레벨"],
        ["icon_key", "string", "아이콘 리소스 키"],
    ]
)

# === TB_Monster (tint_color 삭제, spawn_start_min→chapter, F1 테마) ===
write_sheet(wb, "TB_Monster",
    ["mon_id", "mon_name", "is_boss", "base_hp", "base_speed", "contact_damage",
     "scale", "chapter", "spawn_weight", "special_ability", "sprite_key"],
    [
        ["MON_001", "타이어 좀비", False, 50, 2.5, 8, 1.0, 1, 10, "None", "spr_mon_tire"],
        ["MON_002", "오일 슬라임", False, 30, 4.0, 6, 0.8, 1, 8, "None", "spr_mon_oil"],
        ["MON_003", "메탈 크러셔", False, 120, 1.5, 15, 1.3, 1, 5, "None", "spr_mon_crusher"],
        ["MON_004", "스파크 러너", False, 40, 5.0, 10, 0.9, 1, 6, "FlameTrail", "spr_mon_spark"],
        ["MON_BOSS_001", "워로드 트럭", True, 2000, 1.0, 30, 2.5, 1, 1, "ThreePhase", "spr_boss_truck"],
    ],
    [
        ["mon_id", "string (PK)", "몬스터 고유 ID"],
        ["mon_name", "string", "몬스터 표시 이름"],
        ["is_boss", "bool", "true=보스 몬스터"],
        ["base_hp", "float", "기본 체력"],
        ["base_speed", "float", "이동속도 (units/sec)"],
        ["contact_damage", "float", "접촉 데미지"],
        ["scale", "float", "스프라이트 크기 배율"],
        ["chapter", "int", "등장 챕터"],
        ["spawn_weight", "int", "스폰 가중치"],
        ["special_ability", "string", "특수 능력. None=없음"],
        ["sprite_key", "string", "스프라이트 리소스 키"],
    ]
)

# === TB_MonsterDrop (special_drop 삭제 → screw_amount 추가) ===
write_sheet(wb, "TB_MonsterDrop",
    ["drop_id", "mon_id", "exp_amount", "gold_amount", "screw_amount"],
    [
        ["DROP_001", "MON_001", 3, 5, 0],
        ["DROP_002", "MON_002", 2, 3, 0],
        ["DROP_003", "MON_003", 6, 10, 1],
        ["DROP_004", "MON_004", 4, 6, 0],
        ["DROP_005", "MON_BOSS_001", 50, 200, 20],
    ],
    [
        ["drop_id", "string (PK)", "드롭 고유 ID"],
        ["mon_id", "string (FK)", "몬스터 ID"],
        ["exp_amount", "int", "처치 시 경험치"],
        ["gold_amount", "int", "처치 시 골드"],
        ["screw_amount", "int", "처치 시 나사. 0=없음"],
    ]
)

# === TB_Wave (wave_group_id로 그룹화, 10분 + 연장 생존 페이즈) ===
wave_data = [
    # WG_CH1: 1챕터 기본 웨이브 그룹
    ["WAVE_001", "WG_CH1", 0, "튜토리얼", "MON_001", 4, 15, 1.0, "시작, 타이어 좀비만"],
    ["WAVE_002", "WG_CH1", 1, "초반", "MON_001", 5, 20, 1.1, None],
    ["WAVE_003", "WG_CH1", 2, "초반", "MON_001", 5, 25, 1.2, "오일 슬라임 합류"],
    ["WAVE_004", "WG_CH1", 2, "초반", "MON_002", 2, 25, 1.2, None],
    ["WAVE_005", "WG_CH1", 3, "중반 진입", "MON_001", 4, 30, 1.3, "메탈 크러셔 합류"],
    ["WAVE_006", "WG_CH1", 3, "중반 진입", "MON_002", 3, 30, 1.3, None],
    ["WAVE_007", "WG_CH1", 3, "중반 진입", "MON_003", 1, 30, 1.3, None],
    ["WAVE_008", "WG_CH1", 4, "중반", "MON_001", 4, 35, 1.4, None],
    ["WAVE_009", "WG_CH1", 4, "중반", "MON_002", 3, 35, 1.4, None],
    ["WAVE_010", "WG_CH1", 4, "중반", "MON_003", 2, 35, 1.4, None],
    ["WAVE_011", "WG_CH1", 5, "중반", "MON_001", 5, 40, 1.5, "스파크 러너 합류"],
    ["WAVE_012", "WG_CH1", 5, "중반", "MON_002", 4, 40, 1.5, None],
    ["WAVE_013", "WG_CH1", 5, "중반", "MON_003", 2, 40, 1.5, None],
    ["WAVE_014", "WG_CH1", 5, "중반", "MON_004", 1, 40, 1.5, None],
    ["WAVE_015", "WG_CH1", 6, "중반 후반", "MON_001", 5, 45, 1.6, None],
    ["WAVE_016", "WG_CH1", 6, "중반 후반", "MON_002", 4, 45, 1.6, None],
    ["WAVE_017", "WG_CH1", 6, "중반 후반", "MON_003", 3, 45, 1.6, None],
    ["WAVE_018", "WG_CH1", 6, "중반 후반", "MON_004", 2, 45, 1.6, None],
    ["WAVE_019", "WG_CH1", 7, "후반", "MON_001", 6, 50, 1.8, "밀도 증가"],
    ["WAVE_020", "WG_CH1", 7, "후반", "MON_002", 5, 50, 1.8, None],
    ["WAVE_021", "WG_CH1", 7, "후반", "MON_003", 3, 50, 1.8, None],
    ["WAVE_022", "WG_CH1", 7, "후반", "MON_004", 3, 50, 1.8, None],
    ["WAVE_023", "WG_CH1", 8, "후반", "MON_001", 6, 55, 2.0, None],
    ["WAVE_024", "WG_CH1", 8, "후반", "MON_002", 5, 55, 2.0, None],
    ["WAVE_025", "WG_CH1", 8, "후반", "MON_003", 4, 55, 2.0, None],
    ["WAVE_026", "WG_CH1", 8, "후반", "MON_004", 3, 55, 2.0, None],
    ["WAVE_027", "WG_CH1", 9, "트로피 러시", "MON_001", 7, 60, 2.2, "트로피 수집 마감"],
    ["WAVE_028", "WG_CH1", 9, "트로피 러시", "MON_002", 5, 60, 2.2, None],
    ["WAVE_029", "WG_CH1", 9, "트로피 러시", "MON_003", 4, 60, 2.2, None],
    ["WAVE_030", "WG_CH1", 9, "트로피 러시", "MON_004", 4, 60, 2.2, None],
    # WG_CH1 연장 생존 페이즈 (10분 이후)
    ["WAVE_EXT_01", "WG_CH1", 10, "연장 1단계", "MON_001", 8, 70, 2.5, "쉬움 강화 - 경험치/골드 UP"],
    ["WAVE_EXT_02", "WG_CH1", 10, "연장 1단계", "MON_002", 6, 70, 2.5, None],
    ["WAVE_EXT_03", "WG_CH1", 10, "연장 1단계", "MON_003", 5, 70, 2.5, None],
    ["WAVE_EXT_04", "WG_CH1", 10, "연장 1단계", "MON_004", 4, 70, 2.5, None],
    ["WAVE_EXT_05", "WG_CH1", 12, "연장 2단계", "MON_001", 10, 80, 3.0, "보통 강화 - 보물상자 등장"],
    ["WAVE_EXT_06", "WG_CH1", 12, "연장 2단계", "MON_002", 8, 80, 3.0, None],
    ["WAVE_EXT_07", "WG_CH1", 12, "연장 2단계", "MON_003", 6, 80, 3.0, None],
    ["WAVE_EXT_08", "WG_CH1", 12, "연장 2단계", "MON_004", 5, 80, 3.0, None],
    ["WAVE_EXT_09", "WG_CH1", 14, "연장 3단계", "MON_001", 12, 100, 4.0, "어려움 강화 - 고보상"],
    ["WAVE_EXT_10", "WG_CH1", 14, "연장 3단계", "MON_002", 10, 100, 4.0, None],
    ["WAVE_EXT_11", "WG_CH1", 14, "연장 3단계", "MON_003", 8, 100, 4.0, None],
    ["WAVE_EXT_12", "WG_CH1", 14, "연장 3단계", "MON_004", 7, 100, 4.0, None],
    ["WAVE_EXT_13", "WG_CH1", 16, "위험", "MON_001", 15, 999, 5.0, "감당 불가 수준"],
    ["WAVE_EXT_14", "WG_CH1", 16, "위험", "MON_002", 12, 999, 5.0, None],
    ["WAVE_EXT_15", "WG_CH1", 16, "위험", "MON_003", 10, 999, 5.0, None],
    ["WAVE_EXT_16", "WG_CH1", 16, "위험", "MON_004", 10, 999, 5.0, None],
]

write_sheet(wb, "TB_Wave",
    ["wave_id", "wave_group_id", "time_min", "phase_name", "mon_id", "spawn_per_30s",
     "max_enemies", "difficulty_scale", "note"],
    wave_data,
    [
        ["wave_id", "string (PK)", "웨이브 고유 ID"],
        ["wave_group_id", "string", "웨이브 그룹 ID. 맵에서 참조 (WG_CH1, WG_CH2 등)"],
        ["time_min", "int", "시작 시간 (분)"],
        ["phase_name", "string", "페이즈 이름"],
        ["mon_id", "string (FK)", "스폰할 몬스터 ID"],
        ["spawn_per_30s", "int", "30초당 스폰 수"],
        ["max_enemies", "int", "동시 최대 적 수"],
        ["difficulty_scale", "float", "난이도 배율"],
        ["note", "string", "비고"],
    ]
)

# === TB_Level (등급 확률 컬럼 삭제 - 등급 시스템 제거) ===
write_sheet(wb, "TB_Level",
    ["level", "required_exp", "required_exp_gap", "difficulty_multiplier", "note"],
    [
        [1, 30, 30, 1.0, "시작 튜토리얼 단계"],
        [2, 70, 40, 1.1, None],
        [3, 120, 50, 1.2, None],
        [4, 180, 60, 1.3, None],
        [5, 250, 70, 1.5, "중반 진입"],
        [6, 330, 80, 1.6, None],
        [7, 420, 90, 1.8, None],
        [8, 520, 100, 2.0, "후반 진입"],
        [9, 630, 110, 2.2, None],
        [10, 750, 120, 2.5, "최대 레벨"],
    ],
    [
        ["level", "int (PK)", "플레이어 레벨"],
        ["required_exp", "int", "누적 필요 경험치"],
        ["required_exp_gap", "int", "이전 레벨 대비 증가량"],
        ["difficulty_multiplier", "float", "난이도 배율"],
        ["note", "string", "비고"],
    ]
)

# === TB_Map (1챕터 F1 경기장만) ===
write_sheet(wb, "TB_Map",
    ["map_id", "map_name", "map_desc", "chapter", "wave_group_id", "bg_sprite_key",
     "grid_size", "special_effect", "unlocked_by_default"],
    [
        ["MAP_001", "F1 서킷", "현재 시대 - F1 경기장", 1, "WG_CH1", "map_f1_circuit",
         4, "None", True],
    ],
    [
        ["map_id", "string (PK)", "맵 고유 ID"],
        ["map_name", "string", "맵 표시 이름"],
        ["map_desc", "string", "맵 설명"],
        ["chapter", "int", "챕터 번호"],
        ["wave_group_id", "string", "웨이브 그룹 ID → TB_Wave.wave_group_id"],
        ["bg_sprite_key", "string", "배경 스프라이트 키"],
        ["grid_size", "int", "타일맵 반복 횟수 (중심 기준 ±N)"],
        ["special_effect", "string", "맵 특수 효과. None=없음"],
        ["unlocked_by_default", "bool", "true=기본 해금"],
    ]
)

wb.save("car_survivor_tables.xlsx")
print("Done! Saved car_survivor_tables.xlsx")
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    data_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["):
            break
        data_rows += 1
    print(f"  {name}: {data_rows} rows")
