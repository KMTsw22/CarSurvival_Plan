"""
TB_Car에 flame_offset 컬럼 추가 + TB_Weapon Flamethrower 행 etc 주석 추가 후 재익스포트

사용법: py add_flame_offset.py
"""
import os, sys, io, msgpack, openpyxl
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

CAR_TYPES = [str, str, str, float, float, float, float, float, str, float, str, int, str, bool, str, float]

def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == bool:
        if isinstance(val, bool): return val
        if isinstance(val, str): return val.lower() in ("true", "1", "yes")
        return bool(val)
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)

    # ── 1) TB_Car: flame_offset 컬럼 추가 ──
    ws_car = wb["TB_Car"]
    col_idx = 16  # 16번째 컬럼 (Key[15])

    # 헤더
    header_cell = ws_car.cell(row=1, column=col_idx)
    header_cell.value = "flame_offset"
    header_cell.comment = Comment("화염방사기 오프셋 거리\n차량 앞 불꽃 위치 (최소 3)", "System")

    # 각 행에 기본값 3 설정
    for row_idx, row in enumerate(ws_car.iter_rows(min_row=2), start=2):
        if row[0].value is None:
            break
        ws_car.cell(row=row_idx, column=col_idx, value=3.0)
        print(f"TB_Car 행 {row_idx} ({row[0].value}): flame_offset = 3.0")

    # ── 2) TB_Weapon: Flamethrower 행 etc 주석 추가 ──
    ws_wpn = wb["TB_Weapon"]

    FLAME_COMMENTS = {
        13: "etc_value1: 화상 데미지\n기본 8",
        14: "etc_value2: 기본 반경\n기본 2",
        15: "etc_value3: 레벨당 반경 증가\n기본 0.3",
        16: "etc_value4: 화상 지속시간 (초)\n기본 3",
        17: "etc_value5: (미사용)",
    }

    # 헤더 주석 업데이트 (Flamethrower 정보 추가)
    HEADER_ADDITIONS = {
        13: "\nFlamethrower: 화상 데미지",
        14: "\nFlamethrower: 기본 반경",
        15: "\nFlamethrower: 레벨당 반경 증가",
        16: "\nFlamethrower: 화상 지속시간 (초)",
    }
    for col, addition in HEADER_ADDITIONS.items():
        cell = ws_wpn.cell(row=1, column=col)
        existing = cell.comment.text if cell.comment else f"etc_value{col - 12}"
        cell.comment = Comment(existing + addition, "System")
        print(f"TB_Weapon 헤더 col {col}: 주석 업데이트")

    # Flamethrower 행 주석
    for row in ws_wpn.iter_rows(min_row=2):
        if row[0].value is None: break
        weapon_type = row[6].value  # weapon_type 컬럼 (0-based index 6)
        if weapon_type == "Flamethrower":
            for col, text in FLAME_COMMENTS.items():
                row[col - 1].comment = Comment(text, "System")
            print(f"{row[0].value} (Flamethrower): 행 주석 추가")

    wb.save(XLSX_PATH)
    wb.close()
    print("\n엑셀 저장 완료!")

    # ── 3) TB_Car 재익스포트 ──
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_Car"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["): break
        rows.append([cast(row[i] if i < len(row) else None, CAR_TYPES[i]) for i in range(len(CAR_TYPES))])
    wb.close()

    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_Car.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)
    print(f"TB_Car.bytes 재생성: {len(rows)} rows, {len(packed)} bytes")

if __name__ == "__main__":
    main()
