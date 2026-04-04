"""
TB_Weapon ChainLightning 행 etc 주석 추가 + 헤더 주석 업데이트

사용법: py update_chain_lightning_comments.py
"""
import os, sys, io, openpyxl
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")

CHAIN_COMMENTS = {
    13: "etc_value1: 기본 타격 인원\n기본 3",
    14: "etc_value2: 레벨당 추가 인원\n기본 1",
    15: "etc_value3: 체인 범위\n기본 5",
    16: "etc_value4: (미사용)",
    17: "etc_value5: (미사용)",
}

HEADER_ADDITIONS = {
    13: "\nChainLightning: 기본 타격 인원",
    14: "\nChainLightning: 레벨당 추가 인원",
    15: "\nChainLightning: 체인 범위",
}

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    # 헤더 주석 업데이트
    for col, addition in HEADER_ADDITIONS.items():
        cell = ws.cell(row=1, column=col)
        existing = cell.comment.text if cell.comment else f"etc_value{col - 12}"
        cell.comment = Comment(existing + addition, "System")
        print(f"헤더 col {col}: 주석 업데이트")

    # ChainLightning 행 주석
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None: break
        weapon_type = row[6].value
        if weapon_type == "ChainLightning":
            for col, text in CHAIN_COMMENTS.items():
                row[col - 1].comment = Comment(text, "System")
            print(f"{row[0].value} (ChainLightning): 행 주석 추가")

    wb.save(XLSX_PATH)
    wb.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
