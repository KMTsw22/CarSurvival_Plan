"""
TB_Weapon EMPPulse 행 etc 주석 업데이트 (상시 자기장 방식)

사용법: py update_emp_comments.py
"""
import os, sys, io, openpyxl
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")

EMP_COMMENTS = {
    13: "etc_value1: 타격 간격 (초)\n기본 0.5",
    14: "etc_value2: 기본 반경\n기본 3",
    15: "etc_value3: 레벨당 반경 증가\n기본 0.5",
    16: "etc_value4: (미사용)",
    17: "etc_value5: (미사용)",
}

HEADER_ADDITIONS = {
    13: "\nEMPPulse: 타격 간격 (초)",
    14: "\nEMPPulse: 기본 반경",
    15: "\nEMPPulse: 레벨당 반경 증가",
}

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    for col, addition in HEADER_ADDITIONS.items():
        cell = ws.cell(row=1, column=col)
        existing = cell.comment.text if cell.comment else f"etc_value{col - 12}"
        cell.comment = Comment(existing + addition, "System")
        print(f"헤더 col {col}: 주석 업데이트")

    for row in ws.iter_rows(min_row=2):
        if row[0].value is None: break
        if row[6].value == "EMPPulse":
            for col, text in EMP_COMMENTS.items():
                row[col - 1].comment = Comment(text, "System")
            print(f"{row[0].value} (EMPPulse): 행 주석 업데이트")

    wb.save(XLSX_PATH)
    wb.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
