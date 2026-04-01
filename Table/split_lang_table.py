"""TB_Lang_LevelUpSelect를 name/des 두 테이블로 분리"""
import openpyxl

wb = openpyxl.load_workbook('car_survivor_tables.xlsx')

# 기존 테이블 읽기
ws_old = wb['TB_Lang_LevelUpSelect']
names = {}
descs = {}
for i, row in enumerate(ws_old.iter_rows(values_only=True)):
    if i == 0: continue
    key, ko, en = row[0], row[1], row[2]
    if key is None: break
    if key.endswith('_NAME'):
        # LANG_WPN_001_NAME -> WPN_001
        item_id = key.replace('LANG_', '').replace('_NAME', '')
        names[item_id] = (ko or '', en or '')
    elif key.endswith('_DESC'):
        item_id = key.replace('LANG_', '').replace('_DESC', '')
        descs[item_id] = (ko or '', en or '')

# 기존 시트 삭제
del wb['TB_Lang_LevelUpSelect']

# === TB_LangLevelUpSelect_name ===
ws_name = wb.create_sheet('TB_LangLevelUpSelect_name')
ws_name.cell(row=1, column=1, value='item_id')
ws_name.cell(row=1, column=2, value='ko')
ws_name.cell(row=1, column=3, value='en')
row = 2
for item_id, (ko, en) in names.items():
    ws_name.cell(row=row, column=1, value=item_id)
    ws_name.cell(row=row, column=2, value=ko)
    ws_name.cell(row=row, column=3, value=en)
    row += 1

# === TB_LangLevelUpSelect_des ===
ws_desc = wb.create_sheet('TB_LangLevelUpSelect_des')
ws_desc.cell(row=1, column=1, value='item_id')
ws_desc.cell(row=1, column=2, value='ko')
ws_desc.cell(row=1, column=3, value='en')
row = 2
for item_id, (ko, en) in descs.items():
    ws_desc.cell(row=row, column=1, value=item_id)
    ws_desc.cell(row=row, column=2, value=ko)
    ws_desc.cell(row=row, column=3, value=en)
    row += 1

wb.save('car_survivor_tables.xlsx')

print('Done!')
print(f'  TB_LangLevelUpSelect_name: {len(names)} entries')
for k, (ko, en) in names.items():
    print(f'    {k} -> ko:{ko} / en:{en}')
print(f'  TB_LangLevelUpSelect_des: {len(descs)} entries')
for k, (ko, en) in descs.items():
    print(f'    {k} -> ko:{ko} / en:{en}')
