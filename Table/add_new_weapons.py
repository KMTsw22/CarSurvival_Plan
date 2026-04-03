"""
TB_Weapon에 새 무기 5종 추가 + damage_per_level 컬럼 추가 + .bytes 재익스포트

사용법: py add_new_weapons.py
주의: 엑셀 파일을 먼저 닫고 실행하세요!
"""
import os, sys, io, msgpack, openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

# 18컬럼 (기존 17 + damage_per_level)
TYPES = [str, str, str, float, str, str, str, float, float, int, int, str,
         float, float, float, float, float, float]

# 기존 무기의 damage_per_level 값
EXISTING_DPL = {
    "WPN_001": 5.0,   # 기관총
    "WPN_002": 1.0,   # 오일슬릭
    "WPN_003": 3.0,   # 회전톱날
}

# 새 무기 데이터 (18컬럼)
NEW_WEAPONS = [
    # 체인라이트닝: etc1=기본타격인원3, etc2=레벨당+1명
    ["WPN_004", "체인 라이트닝", "Sub", 25, "연쇄 번개 공격",
     "Auto", "ChainLightning", 1.5, 0, 5, 8, "ico_chainlightning",
     3, 1, 0, 0, 0, 5],
    # EMP펄스: etc1=기본지속시간2초, etc2=기본반경3, etc3=레벨당반경+0.5
    ["WPN_005", "EMP 펄스", "Sub", 10, "범위 전자기 충격",
     "Auto", "EMPPulse", 6, 2, 5, 8, "ico_emppulse",
     2, 3, 0.5, 0, 0, 3],
    # 화염방사기: etc1=기본지속시간3초, etc2=기본반경2, etc3=레벨당반경+0.3
    ["WPN_006", "화염방사기", "Sub", 8, "전방 화염 방사",
     "Auto", "Flamethrower", 4, 3, 5, 8, "ico_flamethrower",
     3, 2, 0.3, 0, 0, 2],
    # 레이저캐논: etc1=기본레이저1개, etc2=레벨당+1개
    ["WPN_007", "레이저 캐논", "Main", 30, "관통 레이저 발사",
     "Manual", "LaserCannon", 2, 0, 5, 8, "ico_lasercannon",
     1, 1, 0, 0, 0, 8],
    # 미사일포드: etc1=기본미사일2개, etc2=레벨당+1개
    ["WPN_008", "미사일 포드", "Sub", 40, "유도 미사일 발사",
     "Auto", "MissilePod", 3, 0, 5, 8, "ico_missilepod",
     2, 1, 0, 0, 0, 10],
]

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: {XLSX_PATH} 파일을 찾을 수 없습니다.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    # --- 1) damage_per_level 헤더 추가 (컬럼 R = 18번째) ---
    col_idx = 18  # 1-based
    header_cell = ws.cell(row=1, column=col_idx)
    if header_cell.value != "damage_per_level":
        header_cell.value = "damage_per_level"
        header_cell.font = Font(bold=True, color='FFFFFF', size=11)
        header_cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = thin_border
        print("  [+] damage_per_level 헤더 추가 (컬럼 R)")

    # --- 2) 기존 무기에 damage_per_level 값 채우기 ---
    existing_ids = set()
    for row in ws.iter_rows(min_row=2):
        cell_id = row[0].value
        if cell_id is None or str(cell_id).startswith("["):
            break
        existing_ids.add(cell_id)
        # damage_per_level 컬럼 (index 17, 1-based col 18)
        dpl_cell = ws.cell(row=row[0].row, column=col_idx)
        if cell_id in EXISTING_DPL and (dpl_cell.value is None or dpl_cell.value == 0):
            dpl_cell.value = EXISTING_DPL[cell_id]
            dpl_cell.border = thin_border
            print(f"  [U] {cell_id}: damage_per_level = {EXISTING_DPL[cell_id]}")

    # --- 3) 새 무기 행 추가 ---
    # 데이터 영역의 마지막 행 찾기
    last_data_row = 1
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None or str(row[0].value).startswith("["):
            break
        last_data_row = row[0].row

    added = 0
    for weapon in NEW_WEAPONS:
        weapon_id = weapon[0]
        if weapon_id in existing_ids:
            # 이미 존재하면 etc 값만 업데이트
            for row in ws.iter_rows(min_row=2):
                if row[0].value == weapon_id:
                    for ci in range(len(weapon)):
                        cell = ws.cell(row=row[0].row, column=ci + 1)
                        cell.value = weapon[ci]
                        cell.border = thin_border
                    print(f"  [U] {weapon_id} 업데이트")
                    break
        else:
            # 새 행 추가
            last_data_row += 1
            for ci in range(len(weapon)):
                cell = ws.cell(row=last_data_row, column=ci + 1)
                cell.value = weapon[ci]
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            added += 1
            print(f"  [+] {weapon_id} ({weapon[1]}) 추가 (행 {last_data_row})")

    print(f"\n총 {added}개 무기 추가됨")

    # 저장
    wb.save(XLSX_PATH)
    wb.close()
    print(f"엑셀 저장 완료: {XLSX_PATH}")

    # --- 4) .bytes 재익스포트 ---
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_Weapon"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["):
            break
        row_data = [cast(row[i] if i < len(row) else None, TYPES[i]) for i in range(len(TYPES))]
        rows.append(row_data)
    wb.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_Weapon.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)
    print(f"TB_Weapon.bytes: {len(rows)} rows, {len(packed)} bytes")


if __name__ == "__main__":
    main()
