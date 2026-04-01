"""
TB_Weapon에서 SawBlade의 icon_key를 ico_spinblade로 수정 후 재익스포트

사용법: py fix_spinblade_icon.py
"""
import os
import sys
import io
import msgpack
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

TYPES = [str, str, str, float, str, str, str, float, float, int, int, str, float, float, float, float, float]

def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val

def main():
    # 1) 엑셀에서 SawBlade의 icon_key 수정
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    updated = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        weapon_type_col = row[6].value  # weapon_type (col index 6)
        if weapon_type_col == "SawBlade":
            icon_col = row[11]  # icon_key (col index 11)
            old_val = icon_col.value
            icon_col.value = "ico_spinblade"
            print(f"행 {row_idx}: icon_key '{old_val}' → 'ico_spinblade'")
            updated = True

    if updated:
        wb.save(XLSX_PATH)
        print("엑셀 저장 완료")
    else:
        print("SawBlade 무기를 찾을 수 없습니다")

    wb.close()

    # 2) TB_Weapon 재익스포트
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_Weapon"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["):
            break
        converted = [cast(row[i] if i < len(row) else None, TYPES[i]) for i in range(len(TYPES))]
        rows.append(converted)

    wb.close()

    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_Weapon.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)

    print(f"TB_Weapon.bytes 재생성 완료: {len(rows)} rows, {len(packed)} bytes")

if __name__ == "__main__":
    main()
