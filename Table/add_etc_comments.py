"""
TB_Weapon 헤더에 etc_value 주석 추가 + 각 행에도 무기별 설명 주석

사용법: py add_etc_comments.py
"""
import os, sys, io, openpyxl
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")

# 헤더 주석 (1-based column index)
HEADER_COMMENTS = {
    13: "etc_value1\n\nMachineGun: (미사용)\nOilSlick: 감속 비율 (%)\nSawBlade: 회전 속도 (도/초)",
    14: "etc_value2\n\nMachineGun: (미사용)\nOilSlick: 감속 지속 시간 (초)\nSawBlade: 궤도 반경",
    15: "etc_value3\n\nMachineGun: (미사용)\nOilSlick: 기본 반경\nSawBlade: 타격 간격 (초)",
    16: "etc_value4 ★레벨당 증가량★\n\nMachineGun: 레벨당 데미지 증가\nOilSlick: 레벨당 범위 증가\nSawBlade: (미사용, 개수=레벨)",
    17: "etc_value5\n\n(예비 — 미사용)",
}

# 무기별 행 주석 (weapon_type → etc1~5 설명)
WEAPON_COMMENTS = {
    "MachineGun": {
        13: "미사용",
        14: "미사용",
        15: "미사용",
        16: "레벨당 DMG 증가량\n예: 20 → Lv1=+20, Lv3=+60",
        17: "미사용",
    },
    "OilSlick": {
        13: "감속 비율 (%)\n예: 50 → 50% 감속",
        14: "감속 지속 시간 (초)",
        15: "기본 반경\n레벨 1일 때 크기",
        16: "레벨당 범위 증가\n예: 0.2 → Lv1=+0.2, Lv3=+0.6",
        17: "미사용",
    },
    "SawBlade": {
        13: "회전 속도 (도/초)\n예: 180 → 초당 반바퀴",
        14: "궤도 반경\n플레이어 주변 도는 거리",
        15: "타격 간격 (초)\n같은 적에 연속히트 쿨타임",
        16: "미사용\n(톱날 개수 = 레벨)",
        17: "미사용",
    },
}

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    # 헤더 주석
    for col, text in HEADER_COMMENTS.items():
        cell = ws.cell(row=1, column=col)
        cell.comment = Comment(text, "System")
        print(f"헤더 col {col}: 주석 추가")

    # 행별 주석
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None: break
        weapon_type = row[6].value
        if weapon_type in WEAPON_COMMENTS:
            for col, text in WEAPON_COMMENTS[weapon_type].items():
                cell = row[col - 1]  # 0-based
                cell.comment = Comment(text, "System")
            print(f"{row[0].value} ({weapon_type}): 행 주석 추가")

    wb.save(XLSX_PATH)
    wb.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
