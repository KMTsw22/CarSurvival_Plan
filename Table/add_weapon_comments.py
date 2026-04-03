"""
TB_Weapon etc_value 셀에 빨간 메모(comment) 추가 + .bytes 재익스포트

사용법: py add_weapon_comments.py
주의: 엑셀 파일을 먼저 닫고 실행하세요!
"""
import os, sys, io, msgpack, openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

TYPES = [str, str, str, float, str, str, str, float, float, int, int, str,
         float, float, float, float, float, float]

# etc_value 컬럼: 1-based col 13~17 (etc1~5), col 18 (damage_per_level)
# 무기별 메모 정의
# key = weapon_id, value = { col: comment_text }
COMMENTS = {
    "WPN_001": {
        4:  "기본 데미지",
        13: "(미사용)",
        14: "(미사용)",
        15: "(미사용)",
        16: "레벨당 데미지 증가 (레거시, R컬럼 사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_002": {
        4:  "초당 데미지 (DPS)",
        13: "감속 비율 (50 = 50%)",
        14: "감속 지속시간 (초)",
        15: "기본 웅덩이 반경",
        16: "레벨당 반경 증가량",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_003": {
        4:  "타격당 데미지",
        13: "회전 속도 (도/초)",
        14: "궤도 반경",
        15: "타격 간격 (초)",
        16: "(미사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_004": {
        4:  "체인 1타당 데미지",
        8:  "쿨타임 (초)",
        13: "기본 타격 인원수",
        14: "레벨당 추가 인원 (+1)",
        15: "(미사용)",
        16: "(미사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_005": {
        4:  "범위 기본 데미지",
        8:  "쿨타임 (초)",
        9:  "스턴 지속시간 (초)",
        13: "기본 스턴 지속시간 (초)",
        14: "기본 반경",
        15: "레벨당 반경 증가량",
        16: "(미사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_006": {
        4:  "초당 데미지 (DPS)",
        8:  "쿨타임 (초)",
        9:  "화염 지속시간 (초)",
        13: "기본 화염 지속시간 (초)",
        14: "기본 화염 반경",
        15: "레벨당 반경 증가량",
        16: "(미사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_007": {
        4:  "레이저 1타당 데미지",
        8:  "쿨타임 (초)",
        13: "기본 레이저 줄기 수",
        14: "레벨당 레이저 추가 (+1)",
        15: "(미사용)",
        16: "(미사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
    "WPN_008": {
        4:  "미사일 1발당 데미지",
        8:  "쿨타임 (초)",
        13: "기본 미사일 수",
        14: "레벨당 미사일 추가 (+1)",
        15: "(미사용)",
        16: "(미사용)",
        17: "(미사용)",
        18: "레벨당 데미지 증가량",
    },
}

AUTHOR = "TableDesigner"


def cast(val, t):
    if val is None:
        return "" if t == str else 0 if t == int else 0.0 if t == float else False
    if t == int: return int(float(val)) if val else 0
    if t == float: return float(val) if val else 0.0
    if t == str: return str(val) if val else ""
    return val


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TB_Weapon"]

    # --- 0) 데이터 영역 셀의 기존 메모 전부 삭제 + 볼드 등 이상한 서식 초기화 ---
    normal_font = Font(size=11)
    for row in ws.iter_rows(min_row=2):
        cell_id = row[0].value
        if cell_id is None or str(cell_id).startswith("["):
            break
        for cell in row:
            cell.comment = None
            # 헤더가 아닌 데이터 행의 폰트를 기본값으로 리셋
            cell.font = normal_font
    print("  [X] 기존 메모 + 서식 초기화")

    # --- 1) 한국어 메모 추가 ---
    count = 0
    for row in ws.iter_rows(min_row=2):
        cell_id = row[0].value
        if cell_id is None or str(cell_id).startswith("["):
            break

        if cell_id not in COMMENTS:
            continue

        weapon_comments = COMMENTS[cell_id]
        for col, text in weapon_comments.items():
            cell = ws.cell(row=row[0].row, column=col)
            cell.comment = Comment(text, AUTHOR)
            cell.comment.width = 200
            cell.comment.height = 40
            count += 1

        print(f"  [C] {cell_id}: {len(weapon_comments)} comments")

    print(f"\n총 {count}개 메모 추가")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"엑셀 저장 완료: {XLSX_PATH}")

    # .bytes 재익스포트
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["TB_Weapon"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).startswith("["):
            break
        row_data = [cast(row[i] if i < len(row) else None, TYPES[i]) for i in range(len(TYPES))]
        rows.append(row_data)
    wb.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_Weapon.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)
    print(f"TB_Weapon.bytes: {len(rows)} rows, {len(packed)} bytes")


if __name__ == "__main__":
    main()
