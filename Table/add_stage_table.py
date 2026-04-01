"""
TB_Stage 시트를 엑셀에 추가하고 .bytes 익스포트

사용법: py add_stage_table.py
"""
import os
import sys
import io
import msgpack
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "car_survivor_tables.xlsx")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "..", "CarSurvior", "Assets", "Resources", "Tables")

# 샘플 스테이지 데이터
# stage_id, map_id, stage_no, boss_mon_id, key_item_id, key_item_count, key_icon, key_name
STAGE_DATA = [
    ["STG_CH1_1", "MAP_CH1", 1, "BOSS_CH1_1", "KEY_CH1_1", 3, "key_red", "붉은 열쇠"],
    ["STG_CH1_2", "MAP_CH1", 2, "BOSS_CH1_2", "KEY_CH1_2", 5, "key_blue", "푸른 열쇠"],
    ["STG_CH1_3", "MAP_CH1", 3, "BOSS_CH1_3", "KEY_CH1_3", 8, "key_gold", "황금 열쇠"],
]

HEADERS = ["stage_id", "map_id", "stage_no", "boss_mon_id", "key_item_id", "key_item_count", "key_icon", "key_name"]

def main():
    # 1) 엑셀에 TB_Stage 시트 추가
    wb = openpyxl.load_workbook(XLSX_PATH)

    if "TB_Stage" in wb.sheetnames:
        del wb["TB_Stage"]
        print("기존 TB_Stage 시트 삭제")

    ws = wb.create_sheet("TB_Stage")

    # 헤더
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    # 데이터
    for row_idx, row_data in enumerate(STAGE_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(XLSX_PATH)
    wb.close()
    print(f"TB_Stage 시트 추가 완료: {XLSX_PATH}")

    # 2) MessagePack .bytes 익스포트
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 타입 변환
    types = [str, str, int, str, str, int, str, str]
    rows = []
    for row_data in STAGE_DATA:
        converted = []
        for val, t in zip(row_data, types):
            if t == int:
                converted.append(int(val))
            elif t == str:
                converted.append(str(val))
            else:
                converted.append(val)
        rows.append(converted)

    packed = msgpack.packb(rows, use_bin_type=True, use_single_float=True)
    out_path = os.path.join(OUTPUT_DIR, "TB_Stage.bytes")
    with open(out_path, "wb") as f:
        f.write(packed)

    print(f"TB_Stage.bytes 생성 완료: {out_path} ({len(rows)} rows, {len(packed)} bytes)")


if __name__ == "__main__":
    main()
