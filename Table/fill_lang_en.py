"""TB_Lang_LevelUpSelect 영어 컬럼 채우기"""
import openpyxl

wb = openpyxl.load_workbook('car_survivor_tables.xlsx')
ws = wb['TB_Lang_LevelUpSelect']

en_texts = {
    'LANG_WPN_001_NAME': 'Machine Gun',
    'LANG_WPN_001_DESC': 'Fires towards cursor direction',
    'LANG_WPN_002_NAME': 'Oil Slick',
    'LANG_WPN_002_DESC': 'Poison puddle (DoT + Slow)',
    'LANG_WPN_003_NAME': 'Saw Blade',
    'LANG_WPN_003_DESC': 'Rotating blade around vehicle',
    'LANG_BOOK_001_NAME': 'Speed Boost',
    'LANG_BOOK_001_DESC': 'Move speed +%',
    'LANG_BOOK_002_NAME': 'Power Up',
    'LANG_BOOK_002_DESC': 'Attack damage +%',
    'LANG_BOOK_003_NAME': 'Vitality',
    'LANG_BOOK_003_DESC': 'Max HP +%',
}

for row in range(2, ws.max_row + 1):
    key = ws.cell(row=row, column=1).value
    if key in en_texts:
        ws.cell(row=row, column=3, value=en_texts[key])

wb.save('car_survivor_tables.xlsx')
print('Done! English texts filled.')
for k, v in en_texts.items():
    print(f'  {k} -> {v}')
